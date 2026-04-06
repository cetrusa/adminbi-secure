import json
import os
import time
import logging
import traceback
from datetime import datetime
from functools import wraps
from typing import Dict, Any, Optional, Callable, TypeVar, List

# RQ Imports
from django_rq import job
from rq import get_current_job

# Project Script Imports
from scripts.extrae_bi.cubo import CuboVentas
from scripts.config import ConfigBasic
from scripts.extrae_bi.cargue_zip import CargueZip
from scripts.extrae_bi.interface import InterfaceContable
from scripts.extrae_bi.interface_siigo import InterfaceContable as InterfaceContableSiigo
from scripts.extrae_bi.matrix import MatrixVentas
from scripts.extrae_bi.plano import InterfacePlano
from scripts.cargue.cargue_infoproducto import ArchivoFuente, CargueInfoProducto
from scripts.cargue.cargue_infoproveedor import CargueInfoVentas
from scripts.cargue.cargue_infoventas_insert import CargueInfoVentasInsert
from scripts.extrae_bi.cargue_maestras import cargar_tablas_maestras, cargar_tabla_individual

# from scripts.StaticPage import StaticPage # No parece usarse
from scripts.extrae_bi.cargue_plano_tsol import CarguePlano
from scripts.extrae_bi.extrae_bi_insert import ExtraeBiConfig, ExtraeBiExtractor
from scripts.extrae_bi.trazabilidad import TrazabilidadExtractor
from apps.home.utils import clean_old_media_files

# Configuración de logging
logger = logging.getLogger(__name__)

from django.conf import settings

# --- Constantes y Tipos (Ajustar según necesidad) ---
# Tomar timeout desde entorno o settings RQ_QUEUES; fallback 7200
DEFAULT_TIMEOUT = int(
    os.getenv(
        "RQ_TASK_TIMEOUT",
        getattr(getattr(settings, "RQ_QUEUES", {}), "get", lambda *_: {})(
            "default", {}
        ).get("DEFAULT_TIMEOUT", 7200),
    )
)
DEFAULT_BATCH_SIZE = 50000
# DEFAULT_RETRY_COUNT = 3 # No usado directamente
# JOB_PROGRESS_KEY_PREFIX = "job_progress_" # No usado directamente
# JOB_META_KEY_PREFIX = "job_meta_" # No usado directamente

# Tipos para tipado
T = TypeVar("T")
ResultDict = Dict[str, Any]


# --- Funciones Helper y Decoradores para RQ ---


def update_job_progress(
    job_id: Optional[str],  # Job ID puede ser None si se llama fuera de contexto
    progress: int,
    status: str = "processing",  # Cambiado default a 'processing'
    meta: Optional[Dict[str, Any]] = None,
) -> None:
    """
    Actualiza el progreso y metadatos de un trabajo RQ en ejecución.
    Intenta obtener el job actual si no se proporciona job_id.
    """
    current_job = get_current_job()
    target_job_id = job_id or (current_job.id if current_job else None)

    if not target_job_id:
        logger.warning(
            "update_job_progress llamado sin job_id y fuera de un contexto de job RQ."
        )
        return

    # Si estamos dentro de un job, usar el objeto job directamente es más eficiente
    job_to_update = (
        current_job if current_job and current_job.id == target_job_id else None
    )

    logger.debug("[update_job_progress] job_id=%s, progress=%s, status=%s, meta=%s", job_id, progress, status, meta)
    if job_to_update:
        if not meta:
            meta = {}
        logger.debug("[update_job_progress] Updating job meta for job_id=%s", target_job_id)
        # Asegurar que 'status' y 'progress' estén en meta para RQ
        # Usar meta.get para evitar sobreescribir valores existentes si no se proporcionan nuevos
        current_meta = job_to_update.meta or {}
        updated_meta = {
            **current_meta,  # Mantener meta existente
            **meta,  # Añadir/Sobreescribir con nuevos meta
            "progress": max(0, min(100, progress)),
            "status": status,
            "updated_at": time.time(),
        }
        job_to_update.meta.update(updated_meta)
        try:
            job_to_update.save_meta()
            logger.debug("[update_job_progress] Meta saved for job_id=%s: progress=%s%% status=%s", target_job_id, updated_meta.get('progress'), status)
            logger.debug(
                f"RQ Job {target_job_id} progress updated: {status} - {updated_meta.get('progress')}%"
            )
        except Exception as e:
            logger.error("[update_job_progress] Error saving meta for job_id=%s: %s", target_job_id, e)
            logger.error(f"Error al guardar meta para RQ Job {target_job_id}: {e}")
    else:
        logger.debug("[update_job_progress] No job found to update for job_id=%s", target_job_id)
        # Si no estamos en el job actual (poco común para progreso), necesitaríamos fetch el job
        # Esto es menos eficiente y generalmente no necesario para updates de progreso
        logger.warning(
            f"Intento de actualizar progreso para Job {target_job_id} fuera de su contexto directo."
        )
        # Podría implementarse fetching el job por ID si es estrictamente necesario


def task_handler(f: Callable[..., T]) -> Callable[..., ResultDict]:
    """
    Decorador que estandariza el manejo de errores y resultados para tareas RQ.
    Proporciona logging, manejo de excepciones, formato de respuesta y tiempo de ejecución.
    """

    @wraps(f)
    def wrapper(*args, **kwargs) -> ResultDict:
        start_time = time.time()
        job = get_current_job()
        task_name = f.__name__
        job_id = job.id if job else "N/A"

        # Inicializa el progreso
        if job:
            update_job_progress(
                job_id, 0, "starting", meta={"stage": "Inicializando tarea"}
            )

        logger.info(
            f"Iniciando tarea RQ {task_name} (Job ID: {job_id}) con args={args}, kwargs={kwargs}"
        )

        try:
            # Ejecuta la función original
            if job:
                update_job_progress(
                    job_id,
                    5,
                    "processing",
                    meta={"stage": "Ejecutando lógica principal"},
                )
            result = f(*args, **kwargs)  # La función decorada debe devolver ResultDict

            # Validar formato del resultado
            if not isinstance(result, dict):
                logger.error(
                    f"Tarea RQ {task_name} (Job ID: {job_id}) devolvió formato incorrecto: {type(result)}"
                )
                result = {
                    "success": False,
                    "error_message": "Formato de resultado interno incorrecto.",
                }

            execution_time = time.time() - start_time
            result["execution_time"] = execution_time  # Añadir tiempo de ejecución

            # Actualizar estado final basado en 'success'
            if result.get("success", False):
                final_stage = result.get("metadata", {}).get(
                    "stage", "Completado"
                )  # Usar stage final si existe
                logger.info(
                    f"Tarea RQ {task_name} (Job ID: {job_id}) completada exitosamente en {execution_time:.2f}s."
                )
                if job:
                    update_job_progress(
                        job_id,
                        100,
                        "completed",
                        meta={"result": result, "stage": final_stage},
                    )
            else:
                final_stage = result.get("metadata", {}).get(
                    "stage", "Fallido"
                )  # Usar stage final si existe
                logger.warning(
                    f"Tarea RQ {task_name} (Job ID: {job_id}) finalizada con error en {execution_time:.2f}s. Mensaje: {result.get('error_message', 'N/A')}"
                )
                if job:
                    update_job_progress(
                        job_id,
                        100,
                        "failed",
                        meta={"result": result, "stage": final_stage},
                    )

            return result

        except Exception as e:
            execution_time = time.time() - start_time
            error_details = traceback.format_exc()
            error_msg = (
                f"Error inesperado en tarea RQ {task_name} (Job ID: {job_id}): {str(e)}"
            )
            logger.error(f"{error_msg}\n{error_details}")

            # Notificar admins por correo sobre la falla
            try:
                from django.core.mail import mail_admins
                mail_admins(
                    subject=f"[DataZenith] Tarea fallida: {task_name}",
                    message=(
                        f"Tarea: {task_name}\n"
                        f"Job ID: {job_id}\n"
                        f"Args: {args}\n"
                        f"DB: {args[0] if args else 'N/A'}\n"
                        f"Error: {e}\n\n"
                        f"{error_details}"
                    ),
                    fail_silently=True,
                )
            except Exception:
                pass

            final_result = {
                "success": False,
                "error_message": error_msg,
                "error_details": error_details,  # Incluir traceback para depuración
                "execution_time": execution_time,
            }
            if job:
                update_job_progress(
                    job_id,
                    100,
                    "failed",
                    meta={"error": str(e), "stage": "Error Crítico"},
                )
            return final_result

    return wrapper


# --- Helpers ---


def _post_process_faltantes_consolidado(file_path):
    """
    Post-procesa el Excel de Faltantes Consolidado (report_id=6).
    Lee el Excel de hoja unica generado por CuboVentas y lo reemplaza
    con un Excel de 3 hojas: Macrozonas, Asesores, Agotados.
    Incluye nombres legibles, fila de totales y formato Excel.
    """
    import pandas as pd
    from openpyxl.styles import Font, numbers, Alignment
    from openpyxl.utils import get_column_letter

    df = pd.read_excel(file_path, engine="openpyxl")

    # Fallback para nombre_producto NULL (LEFT JOIN puede dejar vacío)
    df["nombre_producto"] = df["nombre_producto"].fillna(df["nbProducto"])

    # --- Hoja 1: VENTA X MACROZONAS (agrupado por sede) ---
    df_macro = (
        df.groupby("sede", sort=True)
        .agg(
            macrozona=("macrozona", "first"),
            total_pedidos=("nbCantidadPedidos", "sum"),
            total_facturado=("nbCantidadFacturadaPedidos", "sum"),
            total_faltantes=("nbCantidadFaltantePedidos", "sum"),
            valor_faltante=("vlFaltante", "sum"),
        )
        .reset_index()
    )
    df_macro["pct_faltante"] = (
        (df_macro["total_faltantes"] / df_macro["total_pedidos"].replace(0, 1) * 100)
        .round(2)
    )
    df_macro.columns = [
        "Sede", "Macrozona", "Cant. Pedidos", "Cant. Facturada",
        "Cant. Faltantes", "Valor Faltante", "% Faltante",
    ]

    # --- Hoja 2: VENTA X ASESORES (agrupado por sede + asesor) ---
    df_asesor = (
        df.groupby(["sede", "asesor"], sort=True)
        .agg(
            total_pedidos=("nbCantidadPedidos", "sum"),
            total_facturado=("nbCantidadFacturadaPedidos", "sum"),
            total_faltantes=("nbCantidadFaltantePedidos", "sum"),
            valor_faltante=("vlFaltante", "sum"),
        )
        .reset_index()
    )
    df_asesor["pct_faltante"] = (
        (df_asesor["total_faltantes"] / df_asesor["total_pedidos"].replace(0, 1) * 100)
        .round(2)
    )
    df_asesor.columns = [
        "Sede", "Asesor", "Cant. Pedidos", "Cant. Facturada",
        "Cant. Faltantes", "Valor Faltante", "% Faltante",
    ]

    # --- Hoja 3: AGOTADOS (productos con faltantes > 0) ---
    df_agotados = (
        df[df["nbCantidadFaltantePedidos"] > 0]
        .groupby(["sede", "asesor", "nbProducto", "nombre_producto"], sort=True)
        .agg(
            cantidad_pedida=("nbCantidadPedidos", "sum"),
            cantidad_facturada=("nbCantidadFacturadaPedidos", "sum"),
            cantidad_faltante=("nbCantidadFaltantePedidos", "sum"),
            valor_faltante=("vlFaltante", "sum"),
        )
        .reset_index()
    )
    df_agotados.columns = [
        "Sede", "Asesor", "Cod. Producto", "Producto",
        "Cant. Pedida", "Cant. Facturada", "Cant. Faltante", "Valor Faltante",
    ]

    # --- Guardar Excel multi-hoja ---
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df_macro.to_excel(writer, sheet_name="VENTA X MACROZONAS", index=False)
        df_asesor.to_excel(writer, sheet_name="VENTA X ASESORES", index=False)
        df_agotados.to_excel(writer, sheet_name="AGOTADOS", index=False)

        # --- Formato Excel ---
        bold_font = Font(bold=True)
        num_fmt = '#,##0'
        money_fmt = '#,##0.00'
        pct_fmt = '0.00"%"'

        for sheet_name, sheet_df in [
            ("VENTA X MACROZONAS", df_macro),
            ("VENTA X ASESORES", df_asesor),
            ("AGOTADOS", df_agotados),
        ]:
            ws = writer.sheets[sheet_name]

            # Headers en negrita
            for cell in ws[1]:
                cell.font = bold_font

            # Auto-ajuste de ancho de columnas
            for col_idx, col_name in enumerate(sheet_df.columns, 1):
                max_len = max(
                    len(str(col_name)),
                    sheet_df[col_name].astype(str).str.len().max() if len(sheet_df) > 0 else 0,
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

            # Formato numerico para columnas de valor y porcentaje
            for col_idx, col_name in enumerate(sheet_df.columns, 1):
                col_letter = get_column_letter(col_idx)
                if "Valor" in col_name:
                    for row in range(2, len(sheet_df) + 3):  # +3 para incluir fila totales
                        cell = ws[f"{col_letter}{row}"]
                        if cell.value is not None:
                            cell.number_format = money_fmt
                elif "%" in col_name:
                    for row in range(2, len(sheet_df) + 3):
                        cell = ws[f"{col_letter}{row}"]
                        if cell.value is not None:
                            cell.number_format = pct_fmt
                elif "Cant" in col_name:
                    for row in range(2, len(sheet_df) + 3):
                        cell = ws[f"{col_letter}{row}"]
                        if cell.value is not None:
                            cell.number_format = num_fmt

            # Fila de totales
            total_row = len(sheet_df) + 2  # +1 header, +1 zero-index
            ws.cell(row=total_row, column=1, value="TOTAL").font = bold_font
            for col_idx, col_name in enumerate(sheet_df.columns, 1):
                if col_name in ("Sede", "Macrozona", "Asesor", "Cod. Producto", "Producto"):
                    continue
                if "%" in col_name:
                    # Recalcular % total
                    if "Cant. Faltantes" in sheet_df.columns:
                        t_falt = sheet_df["Cant. Faltantes"].sum()
                        t_ped = sheet_df["Cant. Pedidos"].sum()
                    elif "Cant. Faltante" in sheet_df.columns:
                        t_falt = sheet_df["Cant. Faltante"].sum()
                        t_ped = sheet_df["Cant. Pedida"].sum()
                    else:
                        continue
                    pct_total = round(t_falt / max(t_ped, 1) * 100, 2)
                    cell = ws.cell(row=total_row, column=col_idx, value=pct_total)
                    cell.font = bold_font
                    cell.number_format = pct_fmt
                else:
                    total_val = sheet_df[col_name].sum()
                    cell = ws.cell(row=total_row, column=col_idx, value=total_val)
                    cell.font = bold_font
                    if "Valor" in col_name:
                        cell.number_format = money_fmt
                    else:
                        cell.number_format = num_fmt

    logger.info(
        f"Faltantes Consolidado: {len(df_macro)} sedes, "
        f"{len(df_asesor)} asesores, {len(df_agotados)} productos agotados"
    )


# --- Tareas RQ ---

from django.db import connection


def _enrich_preventa_with_macrozona(result_data, database_name, user_id):
    """Agrega macrozona_id al preview de preventa (report_id=5)."""
    from scripts.conexion import Conexion as con
    from sqlalchemy import text as sa_text

    preview_sample = result_data.get("preview_sample", [])
    preview_headers = result_data.get("preview_headers", [])
    if not preview_sample or "zona_id" not in preview_headers:
        return

    config = ConfigBasic(database_name, user_id).config
    engine = con.ConexionMariadb3(
        str(config["nmUsrIn"]), str(config["txPassIn"]),
        str(config["hostServerIn"]), int(config["portServerIn"]),
        str(config["dbBi"]),
    )
    with engine.connect() as conn:
        rows = conn.execute(
            sa_text("SELECT zona_id, macrozona_id FROM zona WHERE macrozona_id IS NOT NULL")
        ).fetchall()
    zona_map = {str(r[0]): str(r[1]) for r in rows}

    for row in preview_sample:
        zona = str(row.get("zona_id", ""))
        row["macrozona_id"] = zona_map.get(zona, "")

    if "macrozona_id" not in preview_headers:
        idx = preview_headers.index("zona_id") + 1
        preview_headers.insert(idx, "macrozona_id")
    logger.info("[cubo_ventas_task] Preventa enriquecida con macrozona_id")


@job(
    "default", timeout=DEFAULT_TIMEOUT, result_ttl=3600
)  # Usar cola 'default' o una específica, resultado se mantiene 1h
@task_handler  # Aplicar decorador estándar
def cubo_ventas_task(
    database_name,
    IdtReporteIni,
    IdtReporteFin,
    user_id,
    report_id,
    batch_size=DEFAULT_BATCH_SIZE,
):
    """
    Tarea RQ para generar el Cubo de Ventas, reportando progreso detallado.
    Optimizada para grandes volúmenes de datos.
    """
    # Cerrar conexión Django antes de iniciar procesamiento pesado
    try:
        connection.close()
    except Exception:
        pass
    job = get_current_job()
    job_id = job.id if job else None
    logger.info(
        f"Iniciando cubo_ventas_task (RQ Job ID: {job_id}) para DB: {database_name}, Periodo: {IdtReporteIni}-{IdtReporteFin}"
    )

    logger.info("[cubo_ventas_task] INICIO: database_name=%s, IdtReporteIni=%s, IdtReporteFin=%s, user_id=%s, report_id=%s, batch_size=%s", database_name, IdtReporteIni, IdtReporteFin, user_id, report_id, batch_size)

    # Estimación inicial de pasos (puede ajustarse en CuboVentas si es necesario)
    # total_steps_estimate = 5 # No usado directamente aquí

    def rq_update_progress(stage, progress_percent, current_rec=None, total_rec=None):
        """Callback para actualizar el estado de la tarea RQ."""
        # Construir meta data
        meta = {
            "stage": stage,
            # 'current_step': current_step, # Podría añadirse si CuboVentas lo reporta
            # 'total_steps': total_steps_estimate,
        }
        if current_rec is not None:
            meta["records_processed"] = current_rec
        if total_rec is not None:
            meta["total_records_estimate"] = total_rec

        # Llamar a la función helper de RQ
        update_job_progress(
            job_id, int(progress_percent), status="processing", meta=meta
        )

    logger.debug("[cubo_ventas_task] Instanciando CuboVentas...")
    # Instanciar y ejecutar la lógica principal, pasando el callback adaptado para RQ
    cubo_processor = CuboVentas(
        database_name,
        IdtReporteIni,
        IdtReporteFin,
        user_id,
        report_id,
        progress_callback=rq_update_progress,  # <-- Pasar callback adaptado
    )

    # Si CuboVentas soporta batch_size, pásalo aquí o configúralo internamente
    if hasattr(cubo_processor, "batch_size"):
        cubo_processor.batch_size = batch_size

    logger.debug("[cubo_ventas_task] Ejecutando run() de CuboVentas...")
    # run() ahora usa el callback internamente y devuelve el resultado final
    # El decorador @task_handler se encargará del manejo de errores y formato final
    result_data = cubo_processor.run()

    # Post-procesamiento: Faltantes Consolidado (report_id=6) → Excel multi-hoja
    if result_data.get("success") and report_id == 6:
        try:
            _post_process_faltantes_consolidado(result_data["file_path"])
            logger.info("[cubo_ventas_task] Post-procesamiento Faltantes Consolidado completado")
        except Exception as e:
            logger.warning(f"Error en post-procesamiento Faltantes Consolidado: {e}")
            logger.error("[cubo_ventas_task] Error post-procesamiento: %s", e)

    # --- Asegurar que el progreso final se reporte correctamente para el frontend ---
    job = get_current_job()
    job_id = job.id if job else None
    # Si la tarea fue exitosa, reportar 100% y status 'completed' en meta
    if result_data.get("success"):
        update_job_progress(
            job_id,
            100,
            status="completed",
            meta={"stage": "Completado", "file_ready": True},
        )
    else:
        update_job_progress(job_id, 100, status="failed", meta={"stage": "Fallido"})

    # Preview ya viene incluida en result_data desde CuboVentas.run()
    # (se extrae antes de la limpieza de SQLite para evitar tabla inexistente)
    if result_data.get("success"):
        if "preview_headers" not in result_data:
            result_data["preview_headers"] = []
        if "preview_sample" not in result_data:
            result_data["preview_sample"] = []

    # Post-procesamiento: Preventa (report_id=5) → enriquecer preview con macrozona
    if result_data.get("success") and report_id == 5:
        try:
            _enrich_preventa_with_macrozona(result_data, database_name, user_id)
        except Exception as e:
            logger.warning(f"Error enriqueciendo preventa con macrozona: {e}")

    logger.info("[cubo_ventas_task] RESULTADO: %s", result_data)

    # Invalidar cache de KPIs para que reflejen datos frescos
    if result_data.get("success"):
        try:
            from django.core.cache import cache
            cache.delete(f"user_cubo_context_{database_name}")
        except Exception:
            pass

    # Cerrar conexión Django después de finalizar procesamiento pesado
    try:
        connection.close()
    except Exception:
        pass

    logger.info("[cubo_ventas_task] FIN")
    # Devolver directamente el resultado de CuboVentas.run()
    # El decorador @task_handler añadirá execution_time y manejará el estado final.
    return result_data

@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def matrix_task(
    database_name,
    IdtReporteIni,
    IdtReporteFin,
    user_id,
    report_id,
    batch_size=DEFAULT_BATCH_SIZE,
):
    """
    Tarea RQ para generar Matrix de Ventas, reportando progreso detallado.
    Optimizada para grandes volúmenes de datos.
    """
    try:
        connection.close()
    except Exception:
        pass
    job = get_current_job()
    job_id = job.id if job else None
    logger.info(
        f"Iniciando matrix_task (RQ Job ID: {job_id}) para DB: {database_name}, Periodo: {IdtReporteIni}-{IdtReporteFin}"
    )
    logger.info("[matrix_task] INICIO: database_name=%s, IdtReporteIni=%s, IdtReporteFin=%s, user_id=%s, report_id=%s, batch_size=%s", database_name, IdtReporteIni, IdtReporteFin, user_id, report_id, batch_size)

    def rq_update_progress(
        stage,
        progress_percent,
        current_rec=None,
        total_rec=None,
        hoja_idx=None,
        total_hojas=None,
    ):
        meta = {"stage": stage}
        if current_rec is not None:
            meta["records_processed"] = current_rec
        if total_rec is not None:
            meta["total_records_estimate"] = total_rec
        if hoja_idx is not None and total_hojas is not None:
            meta["hoja_actual"] = hoja_idx
            meta["total_hojas"] = total_hojas
            global_percent = int((hoja_idx / total_hojas) * 100)
        else:
            global_percent = progress_percent
        logger.debug("[matrix_task][progreso] stage=%s, hoja_idx=%s, total_hojas=%s, global_percent=%s", stage, hoja_idx, total_hojas, global_percent)
        update_job_progress(job_id, int(global_percent), status="processing", meta=meta)

    logger.debug("[matrix_task] Instanciando MatrixVentas...")
    # Instanciar y ejecutar la lógica principal, pasando el callback adaptado para RQ
    matrix_processor = MatrixVentas(
        database_name,
        IdtReporteIni,
        IdtReporteFin,
        user_id,
        report_id,
        progress_callback=rq_update_progress,  # <-- Pasar callback adaptado
    )

    # Si matrix soporta batch_size, pásalo aquí o configúralo internamente
    if hasattr(matrix_processor, "batch_size"):
        matrix_processor.batch_size = batch_size

    logger.debug("[matrix_task] Ejecutando run() de MatrixVentas...")
    # run() ahora usa el callback internamente y devuelve el resultado final
    # El decorador @task_handler se encargará del manejo de errores y formato final
    result_data = (
        matrix_processor.run()
    )  # batch_size se pasa en __init__ o se usa default

    logger.info("[matrix_task] RESULTADO: %s", result_data)

    # Cerrar conexión Django después de finalizar procesamiento pesado
    try:
        connection.close()
    except Exception:
        pass

    logger.info("[matrix_task] FIN")
    # Devolver directamente el resultado de CuboVentas.run()
    # El decorador @task_handler añadirá execution_time y manejará el estado final.
    return result_data



@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def interface_task(
    database_name,
    IdtReporteIni,
    IdtReporteFin,
    user_id,
    report_id,
    batch_size=DEFAULT_BATCH_SIZE,
):
    """
    Tarea RQ para generar Interface Contable, reportando progreso detallado.
    Optimizada para grandes volúmenes de datos.
    """
    try:
        connection.close()
    except Exception:
        pass
    job = get_current_job()
    job_id = job.id if job else None
    logger.info(
        f"Iniciando interface_task (RQ Job ID: {job_id}) para DB: {database_name}, Periodo: {IdtReporteIni}-{IdtReporteFin}"
    )
    logger.info("[interface_task] INICIO: database_name=%s, IdtReporteIni=%s, IdtReporteFin=%s, user_id=%s, report_id=%s, batch_size=%s", database_name, IdtReporteIni, IdtReporteFin, user_id, report_id, batch_size)

    def rq_update_progress(
        stage,
        progress_percent,
        current_rec=None,
        total_rec=None,
        hoja_idx=None,
        total_hojas=None,
    ):
        meta = {"stage": stage}
        if current_rec is not None:
            meta["records_processed"] = current_rec
        if total_rec is not None:
            meta["total_records_estimate"] = total_rec
        if hoja_idx is not None and total_hojas is not None:
            meta["hoja_actual"] = hoja_idx
            meta["total_hojas"] = total_hojas
            global_percent = int((hoja_idx / total_hojas) * 100)
        else:
            global_percent = progress_percent
        logger.debug("[interface_task][progreso] stage=%s, hoja_idx=%s, total_hojas=%s, global_percent=%s", stage, hoja_idx, total_hojas, global_percent)
        update_job_progress(job_id, int(global_percent), status="processing", meta=meta)

    logger.debug("[interface_task] Instanciando InterfaceContable...")
    # Instanciar y ejecutar la lógica principal, pasando el callback adaptado para RQ
    interface_processor = InterfaceContable(
        database_name,
        IdtReporteIni,
        IdtReporteFin,
        user_id,
        report_id,
        progress_callback=rq_update_progress,  # <-- Pasar callback adaptado
    )

    # Si interface soporta batch_size, pásalo aquí o configúralo internamente
    if hasattr(interface_processor, "batch_size"):
        interface_processor.batch_size = batch_size

    logger.debug("[interface_task] Ejecutando run() de InterfaceContable...")
    # run() ahora usa el callback internamente y devuelve el resultado final
    # El decorador @task_handler se encargará del manejo de errores y formato final
    result_data = (
        interface_processor.run()
    )  # batch_size se pasa en __init__ o se usa default

    logger.info("[interface_task] RESULTADO: %s", result_data)

    # Cerrar conexión Django después de finalizar procesamiento pesado
    try:
        connection.close()
    except Exception:
        pass

    logger.info("[interface_task] FIN")
    # Devolver directamente el resultado de CuboVentas.run()
    # El decorador @task_handler añadirá execution_time y manejará el estado final.
    return result_data


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def interface_siigo_task(
    database_name,
    IdtReporteIni,
    IdtReporteFin,
    user_id,
    report_id,
    batch_size=DEFAULT_BATCH_SIZE,
):
    """
    Tarea RQ para generar Interface Contable con formato SIIGO.
    """
    try:
        connection.close()
    except Exception:
        pass

    job = get_current_job()
    job_id = job.id if job else None
    logger.info(
        f"Iniciando interface_siigo_task (RQ Job ID: {job_id}) para DB: {database_name}, Periodo: {IdtReporteIni}-{IdtReporteFin}"
    )
    logger.info("[interface_siigo_task] INICIO: database_name=%s, IdtReporteIni=%s, IdtReporteFin=%s, user_id=%s, report_id=%s, batch_size=%s", database_name, IdtReporteIni, IdtReporteFin, user_id, report_id, batch_size)

    def rq_update_progress(
        stage,
        progress_percent,
        current_rec=None,
        total_rec=None,
        hoja_idx=None,
        total_hojas=None,
    ):
        meta = {"stage": stage}
        if current_rec is not None:
            meta["records_processed"] = current_rec
        if total_rec is not None:
            meta["total_records_estimate"] = total_rec
        if hoja_idx is not None and total_hojas is not None:
            meta["hoja_actual"] = hoja_idx
            meta["total_hojas"] = total_hojas
            global_percent = int((hoja_idx / total_hojas) * 100)
        else:
            global_percent = progress_percent
        logger.debug("[interface_siigo_task][progreso] stage=%s, hoja_idx=%s, total_hojas=%s, global_percent=%s", stage, hoja_idx, total_hojas, global_percent)
        update_job_progress(job_id, int(global_percent), status="processing", meta=meta)

    logger.debug("[interface_siigo_task] Instanciando InterfaceContableSiigo...")
    interface_processor = InterfaceContableSiigo(
        database_name,
        IdtReporteIni,
        IdtReporteFin,
        user_id,
        report_id,
        progress_callback=rq_update_progress,
    )

    if hasattr(interface_processor, "batch_size"):
        interface_processor.batch_size = batch_size

    logger.debug("[interface_siigo_task] Ejecutando run() de InterfaceContableSiigo...")
    result_data = interface_processor.run()

    logger.info("[interface_siigo_task] RESULTADO: %s", result_data)

    try:
        connection.close()
    except Exception:
        pass

    logger.info("[interface_siigo_task] FIN")
    return result_data


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def plano_task(
    database_name,
    IdtReporteIni,
    IdtReporteFin,
    user_id,
    report_id,
    batch_size=DEFAULT_BATCH_SIZE,
):
    """
    Tarea RQ: Genera archivos planos a partir de datos (InterfacePlano).
    """
    job = get_current_job()
    job_id = job.id if job else None

    logger.info("[plano_task] INICIO: database_name=%s, IdtReporteIni=%s, IdtReporteFin=%s, user_id=%s, report_id=%s, batch_size=%s", database_name, IdtReporteIni, IdtReporteFin, user_id, report_id, batch_size)
    logger.debug("[plano_task] Working directory: %s", os.getcwd())
    logger.debug("[plano_task] media/mydata.db exists? %s", os.path.exists(os.path.join('media', 'mydata.db')))
    logger.debug("[plano_task] media/ dir exists? %s", os.path.exists('media'))
    logger.debug("[plano_task] User: %s", os.environ.get('USERNAME') or os.environ.get('USER'))

    # Callback robusto y uniforme para progreso
    def rq_update_progress(
        stage,
        progress_percent,
        current_rec=None,
        total_rec=None,
        hoja_idx=None,
        total_hojas=None,
        status=None,
        meta=None,
        **kwargs,
    ):
        meta_dict = {"stage": stage}
        if current_rec is not None:
            meta_dict["records_processed"] = current_rec
        if total_rec is not None:
            meta_dict["total_records_estimate"] = total_rec
        if hoja_idx is not None and total_hojas is not None:
            meta_dict["hoja_actual"] = hoja_idx
            meta_dict["total_hojas"] = total_hojas
            global_percent = int((hoja_idx / total_hojas) * 100)
        else:
            global_percent = progress_percent
        if status is not None:
            meta_dict["status"] = status
        if meta is not None:
            meta_dict.update(meta)
        logger.debug("[plano_task][progreso] stage=%s, hoja_idx=%s, total_hojas=%s, global_percent=%s, status=%s, meta=%s", stage, hoja_idx, total_hojas, global_percent, status, meta)
        update_job_progress(
            job_id, int(global_percent), status=(status or "processing"), meta=meta_dict
        )

    logger.debug("[plano_task] Instanciando InterfacePlano...")
    update_job_progress(job_id, 10, meta={"stage": "Iniciando InterfacePlano"})
    interface = InterfacePlano(
        database_name,
        IdtReporteIni,
        IdtReporteFin,
        user_id,
        report_id,
        progress_callback=rq_update_progress,
    )
    logger.debug("[plano_task] Ejecutando run() de InterfacePlano...")
    update_job_progress(
        job_id, 30, meta={"stage": "Evaluando y procesando datos para plano"}
    )
    resultado = interface.run()
    logger.info("[plano_task] RESULTADO: %s", resultado)

    # --- Asegurar que el resultado siempre tenga 'metadata' relevante ---
    if "metadata" not in resultado or not isinstance(resultado.get("metadata"), dict):
        # Intentar obtener info relevante de InterfacePlano si existe
        total_hojas = None
        hojas_con_datos = None
        if hasattr(interface, "config"):
            hojas1 = getattr(interface, "_obtener_lista_hojas", lambda x: [])(
                "txProcedureCsv"
            )
            hojas2 = getattr(interface, "_obtener_lista_hojas", lambda x: [])(
                "txProcedureCsv2"
            )
            total_hojas = len(hojas1) if hojas1 else len(hojas2)
        # Si el resultado tiene éxito, estimar hojas_con_datos como 1 (mínimo) si no hay info
        if resultado.get("success"):
            hojas_con_datos = 1
        else:
            hojas_con_datos = 0
        resultado["metadata"] = {
            "total_hojas": total_hojas,
            "hojas_con_datos": hojas_con_datos,
        }

    # Reportar progreso final y estado global según éxito o error
    if not resultado.get("success", True):
        update_job_progress(
            job_id,
            100,
            status="failed",
            meta={"stage": "Finalizado con error", "result": resultado},
        )
    else:
        update_job_progress(
            job_id,
            100,
            status="completed",
            meta={"stage": "Finalizado", "result": resultado},
        )
    logger.info("[plano_task] FIN")
    return resultado


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def cargue_zip_task(database_name: str, zip_file_path: str) -> ResultDict:
    """
    Tarea RQ: Procesa un archivo ZIP para carga de datos.
    """
    job = get_current_job()
    job_id = job.id if job else None

    logger.info("[cargue_zip_task] INICIO: database_name=%s, zip_file_path=%s", database_name, zip_file_path)

    # Validar que el archivo exista ANTES de llamar a la lógica principal
    if not os.path.exists(zip_file_path):
        logger.error("[cargue_zip_task] Archivo ZIP no encontrado: %s", zip_file_path)
        logger.error(f"Archivo ZIP no encontrado en cargue_zip_task: {zip_file_path}")
        # Devolver error directamente, el decorador lo manejará
        return {
            "success": False,
            "error_message": f"El archivo ZIP no existe en la ruta: {zip_file_path}",
        }

    logger.debug("[cargue_zip_task] Instanciando CargueZip...")
    update_job_progress(job_id, 10, meta={"stage": "Iniciando CargueZip"})

    cargue_zip = CargueZip(database_name, zip_file_path)
    logger.debug("[cargue_zip_task] Ejecutando procesar_zip()...")
    update_job_progress(job_id, 30, meta={"stage": "Procesando archivo ZIP"})

    # Asume que procesar_zip devuelve ResultDict o puede fallar
    resultado = cargue_zip.procesar_zip()
    logger.info("[cargue_zip_task] RESULTADO: %s", resultado)
    update_job_progress(job_id, 90, meta={"stage": "Finalizando procesamiento ZIP"})
    logger.info("[cargue_zip_task] FIN")
    return resultado


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def cargue_plano_task(database_name: str) -> ResultDict:
    """
    Tarea RQ: Procesa archivos planos para cargar datos (TSOL).
    """
    job = get_current_job()
    job_id = job.id if job else None

    logger.info("[cargue_plano_task] INICIO: database_name=%s", database_name)

    logger.debug("[cargue_plano_task] Instanciando CarguePlano...")
    update_job_progress(job_id, 10, meta={"stage": "Iniciando CarguePlano"})
    cargue_plano = CarguePlano(database_name)  # Asume CarguePlano es para TSOL
    logger.debug("[cargue_plano_task] Ejecutando procesar_plano()...")
    update_job_progress(job_id, 30, meta={"stage": "Procesando archivos planos"})

    # Asume que procesar_plano devuelve ResultDict o puede fallar
    resultado = cargue_plano.procesar_plano()
    logger.info("[cargue_plano_task] RESULTADO: %s", resultado)
    update_job_progress(job_id, 90, meta={"stage": "Finalizando carga de planos"})
    logger.info("[cargue_plano_task] FIN")
    return resultado


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def extrae_bi_task(
    database_name: str,
    IdtReporteIni: str,
    IdtReporteFin: str,
    user_id: Optional[int] = None,
    id_reporte: Optional[int] = None,
    batch_size: int = DEFAULT_BATCH_SIZE,
):
    """
    Tarea RQ: Ejecuta la extracción y procesamiento de datos BI (Extrae_Bi).
    """
    job = get_current_job()
    job_id = job.id if job else None

    logger.info("[extrae_bi_task] INICIO: database_name=%s, IdtReporteIni=%s, IdtReporteFin=%s, user_id=%s, id_reporte=%s, batch_size=%s", database_name, IdtReporteIni, IdtReporteFin, user_id, id_reporte, batch_size)

    def rq_update_progress(meta_dict, progress_percent):
        # meta_dict contiene: stage, tabla, nmReporte, progress
        update_job_progress(job_id, int(progress_percent), meta=meta_dict)

    logger.debug("[extrae_bi_task] Instanciando ExtraeBiConfig y ExtraeBiExtractor...")
    update_job_progress(job_id, 5, meta={"stage": "Iniciando Extrae_Bi"})
    logger.info(
        f"Iniciando extrae_bi_task (RQ Job: {job_id}) para {database_name}, Periodo: {IdtReporteIni}-{IdtReporteFin}, user_id={user_id}, id_reporte={id_reporte}, batch_size={batch_size}"
    )
    config = ExtraeBiConfig(database_name)
    extractor = ExtraeBiExtractor(
        config,
        IdtReporteIni,
        IdtReporteFin,
        user_id=user_id,
        id_reporte=id_reporte,
        batch_size=batch_size,
        progress_callback=rq_update_progress,
    )
    logger.debug("[extrae_bi_task] Ejecutando run() de ExtraeBiExtractor...")
    update_job_progress(job_id, 15, meta={"stage": "Ejecutando extractor principal"})
    result = extractor.run()
    logger.info("[extrae_bi_task] RESULTADO: %s", result)

    # Invalidar cache de KPIs tras actualizar datos BI
    if result.get("success"):
        try:
            from django.core.cache import cache
            cache.delete(f"user_cubo_context_{database_name}")
        except Exception:
            pass

    update_job_progress(job_id, 95, meta={"stage": "Finalizando extracción BI"})
    logger.info("[extrae_bi_task] FIN")
    return result


def clean_media_periodic(hours=4):
    """
    Tarea periódica para limpiar archivos viejos en media/.
    """
    removed = clean_old_media_files(hours=hours)
    logger.info(f"[clean_media_periodic] Archivos eliminados: {removed}")
    return removed


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def cargue_infoventas_task(
    temp_path, database_name, IdtReporteIni, IdtReporteFin, user_id=None
):
    """
    Tarea RQ para el cargue masivo de ventas, usando la clase CargueInfoVentasInsert.
    El archivo temporal se elimina al finalizar.
    """
    job = get_current_job()
    job_id = job.id if job else None

    logger.info("[cargue_infoventas_task] INICIO: temp_path=%s, database_name=%s, IdtReporteIni=%s, IdtReporteFin=%s, user_id=%s, job_id=%s", temp_path, database_name, IdtReporteIni, IdtReporteFin, user_id, job_id)

    def rq_update_progress(percent):
        if job:
            update_job_progress(
                job_id,
                percent,
                status="processing",
                meta={"stage": f"Cargue {percent}%"},
            )
        logger.debug("[cargue_infoventas_task] Progreso: %s%% (job_id=%s)", percent, job_id)

    errores = []
    try:
        logger.debug("[cargue_infoventas_task] Instanciando CargueInfoVentasInsert con: temp_path=%s, database_name=%s, IdtReporteIni=%s, IdtReporteFin=%s, user_id=%s", temp_path, database_name, IdtReporteIni, IdtReporteFin, user_id)
        cargador = CargueInfoVentasInsert(
            temp_path, database_name, IdtReporteIni, IdtReporteFin, user_id=user_id
        )
        try:
            logger.debug("[cargue_infoventas_task] Llamando a procesar_cargue...")
            cargador.procesar_cargue(progress_callback=rq_update_progress)
        except Exception as e:
            # Si ocurre un error parcial, lo guardamos pero seguimos
            error_msg = f"Error parcial durante el proceso de carga: {str(e)}"
            logger.error("[cargue_infoventas_task] Error: %s", error_msg)
            errores.append(error_msg)
        resultado = {
            "success": len(errores) == 0,
            "message": (
                "Carga completada con advertencias."
                if errores
                else "Carga completada exitosamente. Revisa los logs para más detalles."
            ),
        }
        if errores:
            resultado["warnings"] = errores
        logger.info(
            f"cargue_infoventas_task (Job ID: {job_id}) completado. Errores: {errores if errores else 'Ninguno'}"
        )
    except Exception as e:
        logger.error("[cargue_infoventas_task] Error critico: %s", e)
        resultado = {"success": False, "error_message": str(e)}
        logger.error(
            f"Error crítico en cargue_infoventas_task (Job ID: {job_id}): {str(e)}",
            exc_info=True,
        )
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
                logger.info(f"Archivo temporal eliminado: {temp_path}")
                logger.info("[cargue_infoventas_task] Archivo temporal eliminado: %s", temp_path)
            except Exception as e:
                logger.warning(
                    f"No se pudo eliminar el archivo temporal {temp_path}: {str(e)}"
                )
                logger.warning("[cargue_infoventas_task] No se pudo eliminar el archivo temporal %s: %s", temp_path, e)

    logger.info("[cargue_infoventas_task] FIN: %s", resultado)
    return resultado


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def cargue_maestras_task(database_name, tablas_seleccionadas=None):
    """
    Tarea RQ para cargar tablas maestras (dimensiones) desde archivos Excel.
    
    Args:
        database_name: Nombre de la base de datos
        tablas_seleccionadas: Lista de tablas específicas a cargar. Si es None, carga todas.
    """
    job = get_current_job()
    job_id = job.id if job else None

    logger.info("[cargue_maestras_task] INICIO: database_name=%s, tablas_seleccionadas=%s, job_id=%s", database_name, tablas_seleccionadas, job_id)

    resultado = {
        "status": "error",
        "message": "",
        "data": {},
        "total_tiempo": 0,
        "job_id": job_id,
        "success": False
    }

    start_time = time.time()

    try:
        # Validar archivos Excel requeridos
        archivos_requeridos = [
            "media/PROVEE-TSOL.xlsx",
            "media/023-COLGATE PALMOLIVE.xlsx", 
            "media/rutero_distrijass_total.xlsx"
        ]
        
        archivos_faltantes = []
        for archivo in archivos_requeridos:
            if not os.path.exists(archivo):
                archivos_faltantes.append(archivo)
        
        if archivos_faltantes:
            raise FileNotFoundError(f"Archivos faltantes: {', '.join(archivos_faltantes)}")

        # Actualizar progreso inicial
        update_job_progress(job_id, 10, "processing", 
                          meta={"stage": "Validando archivos Excel"})

        # Cargar tablas
        if tablas_seleccionadas:
            # Carga individual de tablas seleccionadas
            logger.info("[cargue_maestras_task] Cargando tablas especificas: %s", tablas_seleccionadas)
            
            resultados_tablas = {}
            total_tablas = len(tablas_seleccionadas)
            
            for i, tabla in enumerate(tablas_seleccionadas):
                try:
                    progreso = 20 + (i * 70 // total_tablas)
                    update_job_progress(job_id, progreso, "processing", 
                                      meta={"stage": f"Cargando tabla: {tabla}"})
                    
                    registros = cargar_tabla_individual(database_name, tabla)
                    resultados_tablas[tabla] = {
                        'status': 'exitoso',
                        'registros': registros
                    }
                    logger.info("[cargue_maestras_task] Tabla %s cargada exitosamente: %s registros", tabla, registros)
                    
                except Exception as e:
                    resultados_tablas[tabla] = {
                        'status': 'error',
                        'error': str(e)
                    }
                    logger.error("[cargue_maestras_task] Error cargando tabla %s: %s", tabla, e)
            
            resultado["data"] = resultados_tablas
        else:
            # Carga completa de todas las tablas
            logger.info("[cargue_maestras_task] Cargando todas las tablas maestras")
            
            update_job_progress(job_id, 20, "processing", 
                              meta={"stage": "Cargando todas las tablas maestras"})
            
            def progress_callback(progreso, mensaje, meta_extra=None):
                meta = {"stage": mensaje}
                if meta_extra:
                    meta.update(meta_extra)
                update_job_progress(job_id, progreso, "processing", meta=meta)
            
            # Importar y usar la clase de cargue directamente para tener control del progreso
            from scripts.extrae_bi.cargue_maestras import CargueTablasMaestras
            try:
                cargador = CargueTablasMaestras(database_name)
                resultados_completos = cargador.cargar_todas_las_tablas(progress_callback)
                resultado["data"] = resultados_completos
            except Exception as e:
                logger.error("[cargue_maestras_task] Error en carga completa: %s", e)
                resultado["data"] = {
                    'status': 'error',
                    'error': str(e),
                    'detalles': getattr(e, 'args', [''])[0]
                }

        # Finalizar
        resultado["total_tiempo"] = time.time() - start_time
        
        # Verificar si hubo errores
        errores = [tabla for tabla, info in resultado["data"].items() 
                  if info.get('status') == 'error']
        exitosos = [tabla for tabla, info in resultado["data"].items() 
                   if info.get('status') == 'exitoso']
        
        if errores:
            update_job_progress(job_id, 100, "completed_with_errors", 
                              meta={"stage": f"Completado con errores: {len(errores)} fallaron, {len(exitosos)} exitosos"})
            resultado["status"] = "completed_with_errors"
            resultado["message"] = f"Proceso completado con errores en {len(errores)} tablas: {', '.join(errores)}"
            resultado["success"] = False
        else:
            update_job_progress(job_id, 100, "completed", 
                              meta={"stage": f"Completado exitosamente: {len(exitosos)} tablas cargadas"})
            resultado["status"] = "success"
            resultado["message"] = f"Todas las tablas cargadas exitosamente: {', '.join(exitosos)}"
            resultado["success"] = True

        logger.info("[cargue_maestras_task] COMPLETADO: %s - %s", resultado['status'], resultado['message'])

    except Exception as e:
        resultado["total_tiempo"] = time.time() - start_time
        resultado["message"] = f"Error en carga de maestras: {str(e)}"
        resultado["success"] = False
        
        update_job_progress(job_id, 100, "failed", 
                          meta={"stage": f"Error: {str(e)}"})
        
        logger.error("[cargue_maestras_task] Error: %s", e)
        logger.error(f"Error en cargue_maestras_task: {str(e)}", exc_info=True)

    logger.info("[cargue_maestras_task] FIN: %s", resultado)
    return resultado


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def cargue_infoproducto_task(
    database_name: str,
    fecha_reporte: str,
    archivos: List[Dict[str, Any]],
):
    """Procesa archivos InfoProducto y los carga a la tabla fact_infoproducto."""

    job = get_current_job()
    job_id = job.id if job else None

    update_job_progress(
        job_id,
        5,
        status="processing",
        meta={"stage": "Validando parámetros InfoProducto"},
    )

    try:
        fecha_obj = datetime.strptime(fecha_reporte, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(
            f"Formato de fecha inválido '{fecha_reporte}'. Se espera AAAA-MM-DD."
        ) from exc

    if not archivos:
        raise ValueError("No se proporcionaron archivos para el cargue de InfoProducto.")

    fuentes: List[ArchivoFuente] = []
    for item in archivos:
        path = item.get("path")
        if not path:
            raise ValueError("Cada archivo enviado debe incluir la ruta 'path'.")
        if not os.path.exists(path):
            raise FileNotFoundError(f"No se encontró el archivo a procesar: {path}")

        fuente_id = item.get("fuente_id")
        if not fuente_id:
            raise ValueError("Cada archivo debe definir 'fuente_id'.")

        fuentes.append(
            ArchivoFuente(
                path=path,
                original_name=item.get("original_name", os.path.basename(path)),
                fuente_id=fuente_id,
                fuente_nombre=item.get(
                    "fuente_nombre", fuente_id.replace("_", " ").title()
                ),
                sede=item.get("sede"),
            )
        )

    if not fuentes:
        raise ValueError("Ninguno de los archivos aportados es válido para el cargue.")

    def progress_callback(percent: int, stage: str, meta: Optional[Dict[str, Any]] = None):
        meta_data = {"stage": stage}
        if meta:
            meta_data.update(meta)
        update_job_progress(
            job_id,
            max(0, min(100, int(percent))),
            status="processing",
            meta=meta_data,
        )

    cargador = CargueInfoProducto(
        database_name=database_name,
        fecha_reporte=fecha_obj,
        progress_callback=progress_callback,
    )

    update_job_progress(
        job_id,
        15,
        status="processing",
        meta={"stage": "Iniciando lectura de archivos InfoProducto"},
    )

    resultado = cargador.cargar_archivos(fuentes)

    update_job_progress(
        job_id,
        100,
        status="completed" if resultado.get("success") else "failed",
        meta={
            "stage": resultado.get("metadata", {}).get(
                "stage", "Carga InfoProducto finalizada"
            ),
            "resultado": resultado,
        },
    )

    return resultado


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler  
def cargue_tabla_individual_task(database_name, nombre_tabla):
    """
    Tarea RQ para cargar una tabla maestra específica.
    
    Args:
        database_name: Nombre de la base de datos
        nombre_tabla: Nombre de la tabla a cargar
    """
    job = get_current_job()
    job_id = job.id if job else None

    logger.info("[cargue_tabla_individual_task] INICIO: database_name=%s, tabla=%s, job_id=%s", database_name, nombre_tabla, job_id)

    resultado = {
        "status": "error",
        "message": "",
        "data": {},
        "job_id": job_id
    }

    start_time = time.time()

    try:
        update_job_progress(job_id, 10, "processing", 
                          meta={"stage": f"Iniciando carga de tabla: {nombre_tabla}"})

        # Validar archivos Excel
        archivos_requeridos = [
            "media/PROVEE-TSOL.xlsx",
            "media/023-COLGATE PALMOLIVE.xlsx", 
            "media/rutero_distrijass_total.xlsx"
        ]
        
        for archivo in archivos_requeridos:
            if not os.path.exists(archivo):
                raise FileNotFoundError(f"Archivo requerido no encontrado: {archivo}")

        update_job_progress(job_id, 25, "processing", 
                          meta={"stage": f"Cargando tabla: {nombre_tabla}"})

        # Cargar tabla específica
        registros = cargar_tabla_individual(database_name, nombre_tabla)
        
        resultado["data"] = {
            nombre_tabla: {
                'status': 'exitoso',
                'registros': registros,
                'tiempo': time.time() - start_time
            }
        }
        
        resultado["status"] = "success"
        resultado["message"] = f"Tabla {nombre_tabla} cargada exitosamente: {registros} registros"
        
        update_job_progress(job_id, 100, "completed", 
                          meta={"stage": f"Completado: {registros} registros cargados"})

        logger.info("[cargue_tabla_individual_task] COMPLETADO: %s - %s registros", nombre_tabla, registros)

    except Exception as e:
        resultado["message"] = f"Error cargando tabla {nombre_tabla}: {str(e)}"
        
        update_job_progress(job_id, 100, "failed", 
                          meta={"stage": f"Error: {str(e)}"})
        
        logger.error("[cargue_tabla_individual_task] Error: %s", e)
        logger.error(f"Error en cargue_tabla_individual_task para {nombre_tabla}: {str(e)}", exc_info=True)

    logger.info("[cargue_tabla_individual_task] FIN: %s", resultado)
    return resultado


# ---------------------------------------------------------------------------
# Tareas de envio de reportes por correo
# ---------------------------------------------------------------------------


def _safe_save_workbook(wb, file_path):
    """Guarda un workbook a un archivo temporal y luego renombra atomicamente.

    Previene corrupcion del archivo original si wb.save() falla a mitad
    de la escritura (ZipFile trunca el destino al abrir en modo 'w').
    """
    import tempfile
    import shutil

    dir_name = os.path.dirname(file_path) or "."
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=dir_name)
    os.close(fd)
    try:
        wb.save(tmp_path)
        shutil.move(tmp_path, file_path)
    except Exception:
        # Limpiar temporal si falla
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise


def _send_report_email(subject, recipients, file_path, body_html):
    """Helper: envia un correo con archivo Excel adjunto."""
    from django.core.mail import EmailMessage

    msg = EmailMessage(
        subject=subject,
        body=body_html,
        to=recipients,
    )
    msg.content_subtype = "html"
    if file_path and os.path.exists(file_path):
        # Validar que el archivo no este vacio/corrupto antes de adjuntar
        file_size = os.path.getsize(file_path)
        if file_size < 500:
            logger.warning(
                "Archivo adjunto sospechosamente pequeno (%d bytes): %s",
                file_size, file_path,
            )
        msg.attach_file(file_path)
    msg.send(fail_silently=False)


def _add_inventario_sheet(file_path, engine, proveedor_ids=None, macrozonas=None):
    """
    Agrega una hoja 'Inventario' a un archivo Excel existente.

    Para proveedores: filtra por idProveedor (solo productos del proveedor).
    Para supervisores: filtra por nbAlmacen inferido de macrozonas (todos los proveedores).
    """
    from openpyxl import load_workbook
    from sqlalchemy import text as sa_text
    from scripts.text_cleaner import TextCleaner

    if not file_path or not os.path.exists(file_path):
        return
    if not file_path.endswith(".xlsx"):
        logger.warning("No se puede agregar hoja Inventario a archivo no-xlsx: %s", file_path)
        return

    # Construir query de inventario
    if proveedor_ids:
        ids_int = [int(x) for x in proveedor_ids]
        placeholders = ", ".join(str(x) for x in ids_int)
        sql = (
            "SELECT i.nbAlmacen, i.nbProducto, p.nmProducto, "
            "p.nmProveedor, i.InvDisponible, "
            "COALESCE(p.flPcioStdaCompra, 0) AS PrecioCosto, "
            "ROUND(i.InvDisponible * COALESCE(p.flPcioStdaCompra, 0), 2) AS ValorInventario "
            "FROM inventario i "
            "JOIN productos p ON p.nbProducto = i.nbProducto "
            f"WHERE p.idProveedor IN ({placeholders}) "
            "ORDER BY i.nbAlmacen, p.nmProducto"
        )
    elif macrozonas:
        macro_int = [int(x) for x in macrozonas]
        macro_placeholders = ", ".join(str(x) for x in macro_int)
        sql = (
            "SELECT i.nbAlmacen, i.nbProducto, p.nmProducto, "
            "p.nmProveedor, p.nmTpCategoria, i.InvDisponible, "
            "COALESCE(p.flPcioStdaCompra, 0) AS PrecioCosto, "
            "ROUND(i.InvDisponible * COALESCE(p.flPcioStdaCompra, 0), 2) AS ValorInventario "
            "FROM inventario i "
            "JOIN productos p ON p.nbProducto = i.nbProducto "
            "WHERE i.nbAlmacen IN ("
            "  SELECT DISTINCT nbAlmacen FROM zona "
            f"  WHERE macrozona_id IN ({macro_placeholders}) "
            "  AND nbAlmacen IS NOT NULL"
            ") "
            "ORDER BY i.nbAlmacen, p.nmProveedor, p.nmProducto"
        )
    else:
        return  # Sin filtro, no agregar inventario

    try:
        with engine.connect() as conn:
            result = conn.execute(sa_text(sql))
            columns = list(result.keys())
            rows = result.fetchall()

        if not rows:
            logger.info("Inventario sin datos para el filtro aplicado, omitiendo hoja.")
            return

        wb = load_workbook(file_path)
        ws = wb.create_sheet(title="Inventario")
        ws.append(columns)
        for row in rows:
            cleaned = tuple(
                TextCleaner.clean_for_excel(v) if isinstance(v, str) else v
                for v in row
            )
            ws.append(cleaned)
        _safe_save_workbook(wb, file_path)
        logger.info("Hoja Inventario agregada: %d filas", len(rows))
    except Exception as exc:
        logger.error("Error agregando hoja Inventario: %s", exc, exc_info=True)


def _add_dashboard_supervisor_sheet(
    file_path, engine, macrozonas, fecha_ini, fecha_fin, database_name,
):
    """
    Agrega hoja 'Dashboard' (primera posición) con KPIs de fuerza de ventas
    por zona.  Cifras en miles de pesos ($K).
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from sqlalchemy import text as sa_text
    from scripts.text_cleaner import TextCleaner

    if not file_path or not os.path.exists(file_path) or not file_path.endswith(".xlsx"):
        return
    if not macrozonas:
        return

    macro_int = [int(x) for x in macrozonas]
    placeholders = ", ".join(str(x) for x in macro_int)

    sql_kpis = (
        "SELECT "
        "  macrozona_id, macro AS macrozona, "
        "  MIN(nbAlmacen) AS bodega, MIN(nbAlmacen) AS almacen, "
        "  nbZona AS zona, nmZona AS nombre_zona, "
        "  SUM(CASE WHEN td = 'FV' THEN vlrAntesIva ELSE 0 END) AS ventas_netas, "
        "  SUM(CASE WHEN td IN ('FD','NC') THEN ABS(vlrAntesIva) ELSE 0 END) AS devoluciones, "
        "  SUM(CASE WHEN td = 'CM' THEN ABS(vlrAntesIva) ELSE 0 END) AS cambios, "
        "  COUNT(DISTINCT CASE WHEN td = 'FV' THEN nbFactura END) AS num_facturas, "
        "  COUNT(DISTINCT CASE WHEN td IN ('FD','NC') THEN nbFactura END) AS fact_dev, "
        "  COUNT(DISTINCT CASE WHEN td = 'CM' THEN nbFactura END) AS fact_cam "
        "FROM cuboventas "
        "WHERE dtContabilizacion BETWEEN :fi AND :ff "
        f"  AND macrozona_id IN ({placeholders}) "
        "GROUP BY macrozona_id, macro, nbZona, nmZona "
        "ORDER BY macrozona_id, nbZona"
    )

    # Impactos: clientes con venta neta > 0 por zona
    sql_impactos = (
        "SELECT zona, COUNT(*) AS impactos FROM ("
        "  SELECT nbZona AS zona, idPuntoVenta "
        "  FROM cuboventas "
        "  WHERE dtContabilizacion BETWEEN :fi AND :ff "
        f"    AND macrozona_id IN ({placeholders}) "
        "  GROUP BY nbZona, idPuntoVenta "
        "  HAVING SUM(vlrAntesIva) > 0"
        ") sub GROUP BY zona"
    )

    sql_faltantes = (
        "SELECT z.zona_id AS zona, COALESCE(SUM(f.vlFaltante), 0) AS faltantes "
        "FROM faltantes f "
        "JOIN zona z ON z.zona_id = f.nbZona "
        "WHERE f.dtContabilizacion BETWEEN :fi AND :ff "
        f"  AND z.macrozona_id IN ({placeholders}) "
        "GROUP BY z.zona_id"
    )

    # Clientes activos por zona (desde tabla rutas)
    sql_cl_activos = (
        "SELECT r.zona_id AS zona, COUNT(DISTINCT r.cliente_id) AS cl_activos "
        "FROM rutas r "
        "JOIN zona z ON z.zona_id = r.zona_id "
        f"WHERE z.macrozona_id IN ({placeholders}) "
        "GROUP BY r.zona_id"
    )

    # Ultima fecha de FV para calcular dias transcurridos reales
    sql_ultima_fv = (
        "SELECT MAX(dtContabilizacion) AS ultima_fv "
        "FROM cuboventas "
        "WHERE td = 'FV' "
        "  AND dtContabilizacion BETWEEN :fi AND :ff "
        f"  AND macrozona_id IN ({placeholders})"
    )

    # Dias habiles del mes (boSeleccionado=0 es habil)
    # Transcurridos se cuentan solo hasta la ultima fecha de FV
    sql_habiles = (
        "SELECT "
        "  SUM(CASE WHEN h.boSeleccionado = 0 THEN 1 ELSE 0 END) AS dias_habiles, "
        "  SUM(CASE WHEN h.boSeleccionado = 0 AND h.dtFecha <= :corte "
        "    THEN 1 ELSE 0 END) AS dias_transcurridos "
        "FROM habiles h "
        "WHERE h.nbMes = MONTH(CURDATE()) "
        "  AND h.nbAnno = YEAR(CURDATE())"
    )

    # Facturas FV con devolución total (DT): contar por zona las FV cuyo
    # nbFactura aparece en un FD/NC con idmotivo DT*, para restarlas del
    # num_facturas.  Se usa IN con subquery materializada: ~0.5 s.
    sql_nf_dt = (
        "SELECT cv.nbZona AS zona, COUNT(DISTINCT cv.nbFactura) AS nf_dt "
        "FROM cuboventas cv "
        "WHERE cv.td = 'FV' "
        "  AND cv.dtContabilizacion BETWEEN :fi AND :ff "
        f"  AND cv.macrozona_id IN ({placeholders}) "
        "  AND cv.nbFactura IN ("
        "    SELECT DISTINCT nbFactura FROM cuboventas "
        "    WHERE td IN ('FD','NC') AND idmotivo LIKE :dt_prefix "
        "    AND dtContabilizacion BETWEEN :fi AND :ff "
        f"    AND macrozona_id IN ({placeholders})"
        "  ) "
        "GROUP BY cv.nbZona"
    )

    try:
        params = {"fi": fecha_ini, "ff": fecha_fin}
        params_dt = {**params, "dt_prefix": "DT%"}
        with engine.connect() as conn:
            rows = conn.execute(sa_text(sql_kpis), params).mappings().all()
            falt_rows = conn.execute(sa_text(sql_faltantes), params).mappings().all()
            imp_rows = conn.execute(sa_text(sql_impactos), params).mappings().all()
            try:
                cl_act_rows = conn.execute(sa_text(sql_cl_activos)).mappings().all()
            except Exception:
                cl_act_rows = []
                logger.info("Tabla rutas no disponible, omitiendo cl. activos.")

            # Facturas FV con devolucion total (DT) por zona
            try:
                nf_dt_rows = conn.execute(sa_text(sql_nf_dt), params_dt).mappings().all()
            except Exception:
                nf_dt_rows = []
                logger.info("No se pudo obtener facturas DT, se omite exclusion.")

            # Dias habiles y proyeccion
            dias_habiles = 0
            dias_transcurridos = 0
            try:
                ufv_row = conn.execute(sa_text(sql_ultima_fv), params).mappings().first()
                fecha_corte = str(ufv_row["ultima_fv"]) if ufv_row and ufv_row["ultima_fv"] else fecha_fin
                dias_row = conn.execute(
                    sa_text(sql_habiles), {"corte": fecha_corte}
                ).mappings().first()
                if dias_row:
                    dias_habiles = int(dias_row["dias_habiles"] or 0)
                    dias_transcurridos = int(dias_row["dias_transcurridos"] or 0)
            except Exception:
                logger.info("Tabla habiles no disponible, omitiendo proyeccion.")

        if not rows:
            logger.info("Dashboard sin datos para macrozonas %s", macrozonas)
            return

        falt_map = {str(r["zona"]): float(r["faltantes"] or 0) for r in falt_rows}
        cl_act_map = {str(r["zona"]): int(r["cl_activos"] or 0) for r in cl_act_rows}
        imp_map = {str(r["zona"]): int(r["impactos"] or 0) for r in imp_rows}
        nf_dt_map = {str(r["zona"]): int(r["nf_dt"] or 0) for r in nf_dt_rows}

        # ── Métricas por zona ──
        zones = []
        tot = {
            "vb": 0, "vn": 0, "dev": 0, "cam": 0, "nf": 0,
            "fd": 0, "fc": 0, "ca": 0, "ci": 0, "falt": 0,
        }
        for r in rows:
            vb = float(r["ventas_netas"] or 0)   # venta bruta (solo FV)
            dev = float(r["devoluciones"] or 0)
            vn = vb - dev                          # venta neta real (FV - FD - NC)
            cam = float(r["cambios"] or 0)
            nf_raw = int(r["num_facturas"] or 0)
            nf = max(nf_raw - nf_dt_map.get(str(r["zona"]), 0), 0)  # excluir DT
            fd = int(r["fact_dev"] or 0)
            fc = int(r["fact_cam"] or 0)
            ci = imp_map.get(str(r["zona"]), 0)
            ca = cl_act_map.get(str(r["zona"]), 0)
            falt = falt_map.get(str(r["zona"]), 0)
            proy = (vn / dias_transcurridos) * dias_habiles if dias_transcurridos > 0 and dias_habiles > 0 else 0

            _tc = TextCleaner.clean_for_excel
            zones.append({
                "zona": _tc(r["zona"]) if isinstance(r["zona"], str) else r["zona"],
                "nombre": _tc(r["nombre_zona"] or ""),
                "macrozona": _tc(r["macrozona"] or ""),
                "almacen": _tc(r["almacen"] or r["bodega"] or ""),
                "vb": vb, "vn": vn, "dev": dev, "cam": cam,
                "nf": nf, "fd": fd, "fc": fc,
                "ca": ca, "ci": ci, "falt": falt, "proy": proy,
                "pct_dev": dev / vb if vb > 0 else 0,
                "pct_cam": cam / vb if vb > 0 else 0,
                "pct_cl": ci / ca if ca > 0 else 0,
                "drop": vn / ci if ci > 0 else 0,
            })
            for k in ("vb", "vn", "dev", "cam", "falt"):
                tot[k] += locals()[k]
            tot["nf"] += nf; tot["fd"] += fd; tot["fc"] += fc
            tot["ca"] += ca; tot["ci"] += ci

        tot["pct_dev"] = tot["dev"] / tot["vb"] if tot["vb"] > 0 else 0
        tot["pct_cam"] = tot["cam"] / tot["vb"] if tot["vb"] > 0 else 0
        tot["pct_cl"] = tot["ci"] / tot["ca"] if tot["ca"] > 0 else 0
        tot["drop"] = tot["vn"] / tot["ci"] if tot["ci"] > 0 else 0
        tot["proy"] = (tot["vn"] / dias_transcurridos) * dias_habiles if dias_transcurridos > 0 and dias_habiles > 0 else 0

        # ── Construir hoja ──
        wb = load_workbook(file_path)
        ws = wb.create_sheet(title="Dashboard", index=0)

        DARK = "2F5496"
        BLUE = "4472C4"
        LIGHT = "D6E4F0"
        RED_C = "C00000"

        fill_dark = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
        fill_blue = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        fill_light = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")
        right_al = Alignment(horizontal="right", vertical="center")
        left_al = Alignment(horizontal="left", vertical="center")
        thin = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )

        # Fila 1-2: título
        NUM_COLS = 17
        for rw in (1, 2):
            ws.merge_cells(start_row=rw, start_column=1, end_row=rw, end_column=NUM_COLS)
            for c in range(1, NUM_COLS + 1):
                ws.cell(row=rw, column=c).fill = fill_dark
        ws["A1"].value = f"Dashboard Fuerza de Ventas — {database_name.upper()}"
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        ws["A1"].alignment = center
        ws["A2"].value = (
            f"Periodo: {fecha_ini} a {fecha_fin}  ·  Cifras en miles de pesos ($K)"
        )
        ws["A2"].font = Font(name="Calibri", size=10, color="FFFFFF")
        ws["A2"].alignment = center

        # Fila 4-5: KPI cards (7 tarjetas × 2 cols)
        lbl_font = Font(name="Calibri", size=9, color="666666")
        val_ok = Font(name="Calibri", size=16, bold=True, color=DARK)
        val_bad = Font(name="Calibri", size=16, bold=True, color=RED_C)
        kpis = [
            ("T. Venta Neta", f"${tot['vn']/1000:,.0f}K", False),
            ("Proyección", f"${tot['proy']/1000:,.0f}K", False),
            ("T. Devolución", f"${tot['dev']/1000:,.0f}K", True),
            ("% Devolución", f"{tot['pct_dev']*100:.1f}%", True),
            ("% Cl. Imp.", f"{tot['pct_cl']*100:.1f}%", False),
            ("T. Cambios", f"${tot['cam']/1000:,.0f}K", True),
            ("Drop Size", f"${tot['drop']/1000:,.0f}K", False),
            ("# Facturas", f"{tot['nf']:,}", False),
        ]
        for i, (label, value, bad) in enumerate(kpis):
            c = i * 2 + 1
            ws.merge_cells(start_row=4, start_column=c, end_row=4, end_column=c + 1)
            ws.merge_cells(start_row=5, start_column=c, end_row=5, end_column=c + 1)
            ws.cell(row=4, column=c, value=label).font = lbl_font
            ws.cell(row=4, column=c).alignment = center
            cv = ws.cell(row=5, column=c, value=value)
            cv.font = val_bad if bad else val_ok
            cv.alignment = center

        # Fila 7: encabezados tabla
        headers = [
            "Zona", "Macrozona", "Almacen",
            "T. Vta Neta ($K)", "Proyección ($K)", "T. Devolución ($K)", "% Dev.",
            "# Fact. Dev.", "T. Cambios ($K)", "% Camb.", "# Fact. Cam.",
            "Cl. Activos", "Cl. Impactados", "% Cl. Imp.", "Drop Size ($K)",
            "# Facturas", "Faltantes ($K)",
        ]
        hdr_font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=7, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = fill_blue
            cell.alignment = center
            cell.border = thin

        # Filas 8+: datos por zona
        pct_cols = (7, 10, 14)  # % Dev., % Camb., % Cl. Imp.
        dat_font = Font(name="Calibri", size=9)
        for ri, z in enumerate(zones, 8):
            vals = [
                z["zona"], z["macrozona"], z["almacen"],
                round(z["vn"] / 1000), round(z["proy"] / 1000),
                round(z["dev"] / 1000), z["pct_dev"],
                z["fd"], round(z["cam"] / 1000), z["pct_cam"], z["fc"],
                z["ca"], z["ci"], z["pct_cl"], round(z["drop"] / 1000),
                z["nf"], round(z["falt"] / 1000),
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = dat_font
                cell.border = thin
                if ci in pct_cols:
                    cell.number_format = "0.0%"
                elif ci >= 4:
                    cell.number_format = "#,##0"
                cell.alignment = left_al if ci <= 3 else right_al
                if ri % 2 == 0:
                    cell.fill = fill_light

        # Fila totales
        tr = 8 + len(zones)
        tot_font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        tot_vals = [
            "TOTAL", "", "",
            round(tot["vn"] / 1000), round(tot["proy"] / 1000),
            round(tot["dev"] / 1000), tot["pct_dev"],
            tot["fd"], round(tot["cam"] / 1000), tot["pct_cam"], tot["fc"],
            tot["ca"], tot["ci"], tot["pct_cl"], round(tot["drop"] / 1000),
            tot["nf"], round(tot["falt"] / 1000),
        ]
        for ci, v in enumerate(tot_vals, 1):
            cell = ws.cell(row=tr, column=ci, value=v)
            cell.font = tot_font
            cell.fill = fill_dark
            cell.border = thin
            if ci in pct_cols:
                cell.number_format = "0.0%"
            elif ci >= 4:
                cell.number_format = "#,##0"
            cell.alignment = left_al if ci <= 3 else right_al

        # Anchos
        widths = [8, 14, 14, 18, 18, 18, 10, 12, 16, 10, 12, 12, 14, 10, 14, 12, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        _safe_save_workbook(wb, file_path)
        logger.info("Hoja Dashboard agregada con %d zonas", len(zones))

    except Exception as exc:
        logger.error("Error agregando hoja Dashboard: %s", exc, exc_info=True)


def _add_devoluciones_dia_sheet(file_path, engine, macrozonas):
    """
    Agrega hoja 'Devoluciones Dia' con el detalle de devoluciones por cliente
    del día actual.  Cifras en miles de pesos ($K).
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from sqlalchemy import text as sa_text
    from scripts.text_cleaner import TextCleaner
    from datetime import date

    if not file_path or not os.path.exists(file_path) or not file_path.endswith(".xlsx"):
        return
    if not macrozonas:
        return

    macro_int = [int(x) for x in macrozonas]
    placeholders = ", ".join(str(x) for x in macro_int)

    sql = (
        "SELECT "
        "  nmZona AS zona, nbDocumento AS documento, "
        "  nmRazonSocial AS cliente, idPuntoVenta AS punto_venta, "
        "  nbFactura AS factura, nbProducto AS cod_producto, "
        "  nmProducto AS producto, "
        "  ABS(cantAsignada) AS cantidad, "
        "  ROUND(ABS(vlrAntesIva) / 1000, 1) AS valor_k, "
        "  motivo "
        "FROM cuboventas "
        "WHERE dtContabilizacion = CURDATE() "
        "  AND td IN ('FD','NC') "
        f"  AND macrozona_id IN ({placeholders}) "
        "ORDER BY nmZona, nmRazonSocial, nbProducto"
    )

    try:
        with engine.connect() as conn:
            result = conn.execute(sa_text(sql))
            rows = result.fetchall()

        if not rows:
            logger.info("Sin devoluciones del dia para macrozonas %s", macrozonas)
            return

        RED_C = "C00000"
        PINK = "FCE4EC"
        fill_red = PatternFill(start_color=RED_C, end_color=RED_C, fill_type="solid")
        fill_pink = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")
        right_al = Alignment(horizontal="right", vertical="center")
        thin = Border(
            left=Side(style="thin", color="F2DCDB"),
            right=Side(style="thin", color="F2DCDB"),
            top=Side(style="thin", color="F2DCDB"),
            bottom=Side(style="thin", color="F2DCDB"),
        )

        wb = load_workbook(file_path)
        ws = wb.create_sheet(title="Devoluciones Dia")

        # Fila 1: título
        ws.merge_cells("A1:J1")
        ws["A1"] = f"Devoluciones del Dia — {date.today().strftime('%d/%m/%Y')}"
        ws["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        ws["A1"].fill = fill_red
        ws["A1"].alignment = center
        for c in range(1, 11):
            ws.cell(row=1, column=c).fill = fill_red

        # Fila 2: resumen
        total_valor = sum(float(r[8]) if r[8] else 0 for r in rows)
        ws.merge_cells("A2:J2")
        ws["A2"] = f"Total: {len(rows)} items  ·  Valor total: ${total_valor:,.1f}K"
        ws["A2"].font = Font(name="Calibri", size=10, bold=True, color=RED_C)
        ws["A2"].alignment = center

        # Fila 3: encabezados
        display_headers = [
            "Zona", "Documento", "Cliente", "Punto Venta", "Factura",
            "Cod. Producto", "Producto", "Cantidad", "Valor ($K)", "Motivo",
        ]
        hdr_font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        for ci, h in enumerate(display_headers, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = fill_red
            cell.alignment = center
            cell.border = thin

        # Datos
        dat_font = Font(name="Calibri", size=9)
        for ri, row in enumerate(rows, 4):
            for ci, v in enumerate(row, 1):
                cleaned = TextCleaner.clean_for_excel(v) if isinstance(v, str) else v
                cell = ws.cell(row=ri, column=ci, value=cleaned)
                cell.font = dat_font
                cell.border = thin
                if ci in (8, 9):
                    cell.alignment = right_al
                    cell.number_format = "#,##0.0" if ci == 9 else "#,##0"
                if ri % 2 == 0:
                    cell.fill = fill_pink

        # Anchos
        widths = [8, 14, 30, 12, 14, 14, 30, 10, 12, 25]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        _safe_save_workbook(wb, file_path)
        logger.info("Hoja Devoluciones Dia agregada: %d filas", len(rows))

    except Exception as exc:
        logger.error("Error agregando hoja Devoluciones Dia: %s", exc, exc_info=True)


def _add_vendedor_proveedor_sheet(
    file_path, engine, macrozonas, fecha_ini, fecha_fin,
):
    """
    Agrega hoja 'Vendedor x Proveedor' con KPIs cruzados:
    venta, proyección, % cumplimiento, clientes impactados,
    % clientes impactados, devoluciones, % devolución.
    Cifras en miles de pesos ($K).
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from sqlalchemy import text as sa_text
    from scripts.text_cleaner import TextCleaner

    if not file_path or not os.path.exists(file_path) or not file_path.endswith(".xlsx"):
        return
    if not macrozonas:
        return

    macro_int = [int(x) for x in macrozonas]
    placeholders = ", ".join(str(x) for x in macro_int)

    # ── Ventas, devoluciones por vendedor × proveedor ──
    sql_main = (
        "SELECT "
        "  cv.nbZona AS zona, cv.nmZona AS vendedor, "
        "  cv.idProveedor AS id_proveedor, cv.nmProveedor AS proveedor, "
        "  SUM(CASE WHEN cv.td = 'FV' THEN cv.vlrAntesIva ELSE 0 END) AS venta, "
        "  SUM(CASE WHEN cv.td IN ('FD','NC') THEN ABS(cv.vlrAntesIva) ELSE 0 END) AS devolucion "
        "FROM cuboventas cv "
        "WHERE cv.dtContabilizacion BETWEEN :fi AND :ff "
        f"  AND cv.macrozona_id IN ({placeholders}) "
        "GROUP BY cv.nbZona, cv.nmZona, cv.idProveedor, cv.nmProveedor "
        "HAVING venta > 0 OR devolucion > 0 "
        "ORDER BY cv.nbZona, cv.nmProveedor"
    )

    # Impactos por vendedor × proveedor: clientes con venta neta > 0
    sql_imp_vp = (
        "SELECT zona, id_proveedor, COUNT(*) AS cl_imp FROM ("
        "  SELECT cv.nbZona AS zona, cv.idProveedor AS id_proveedor, cv.idPuntoVenta "
        "  FROM cuboventas cv "
        "  WHERE cv.dtContabilizacion BETWEEN :fi AND :ff "
        f"    AND cv.macrozona_id IN ({placeholders}) "
        "  GROUP BY cv.nbZona, cv.idProveedor, cv.idPuntoVenta "
        "  HAVING SUM(cv.vlrAntesIva) > 0"
        ") sub GROUP BY zona, id_proveedor"
    )

    # ── Clientes activos por zona (desde tabla rutas) ──
    sql_cl = (
        "SELECT r.zona_id AS zona, COUNT(DISTINCT r.cliente_id) AS total_cl "
        "FROM rutas r "
        "JOIN zona z ON z.zona_id = r.zona_id "
        f"WHERE z.macrozona_id IN ({placeholders}) "
        "GROUP BY r.zona_id"
    )

    # ── Presupuesto/proyección por zona × proveedor (tabla opcional) ──
    sql_pres = (
        "SELECT nbZona, CAST(idProveedor AS CHAR) AS id_prov, "
        "  SUM(vlrPresupuesto) AS proyeccion "
        "FROM presupuesto "
        "WHERE dtPresupuesto BETWEEN :fi AND :ff "
        "GROUP BY nbZona, idProveedor"
    )

    try:
        params = {"fi": fecha_ini, "ff": fecha_fin}
        with engine.connect() as conn:
            rows = conn.execute(sa_text(sql_main), params).mappings().all()
            imp_vp_rows = conn.execute(sa_text(sql_imp_vp), params).mappings().all()
            try:
                cl_rows = conn.execute(sa_text(sql_cl)).mappings().all()
            except Exception:
                cl_rows = []
                logger.info("Tabla rutas no disponible, omitiendo cl. activos.")

        if not rows:
            logger.info("Vendedor×Proveedor sin datos para macrozonas %s", macrozonas)
            return

        cl_map = {str(r["zona"]): int(r["total_cl"] or 0) for r in cl_rows}
        imp_vp_map = {
            (str(r["zona"]), str(r["id_proveedor"])): int(r["cl_imp"] or 0)
            for r in imp_vp_rows
        }

        # Presupuesto (puede no existir en todas las BD)
        pres_map = {}
        try:
            with engine.connect() as conn:
                pres_rows = conn.execute(sa_text(sql_pres), params).mappings().all()
                for r in pres_rows:
                    key = (str(r["nbZona"]), str(r["id_prov"]))
                    pres_map[key] = float(r["proyeccion"] or 0)
        except Exception:
            logger.info("Tabla presupuesto no disponible, omitiendo proyeccion.")

        # ── Construir datos ──
        data = []
        for r in rows:
            venta_bruta = float(r["venta"] or 0)
            dev = float(r["devolucion"] or 0)
            venta_neta = venta_bruta - dev
            cl_imp = imp_vp_map.get((str(r["zona"]), str(r["id_proveedor"])), 0)
            total_cl = cl_map.get(str(r["zona"]), 0)
            proy = pres_map.get((str(r["zona"]), str(r["id_proveedor"])), 0)

            _tc = TextCleaner.clean_for_excel
            data.append({
                "zona": _tc(r["zona"]) if isinstance(r["zona"], str) else r["zona"],
                "vendedor": _tc(r["vendedor"] or ""),
                "proveedor": _tc(r["proveedor"] or ""),
                "venta_neta": venta_neta,
                "proyeccion": proy,
                "pct_cumpl": venta_neta / proy if proy > 0 else 0,
                "cl_imp": cl_imp,
                "pct_cl": cl_imp / total_cl if total_cl > 0 else 0,
                "devolucion": dev,
                "pct_dev": dev / venta_bruta if venta_bruta > 0 else 0,
            })

        # ── Construir hoja ──
        wb = load_workbook(file_path)
        ws = wb.create_sheet(title="Vendedor x Proveedor")

        DARK = "2F5496"
        BLUE = "4472C4"
        LIGHT = "D6E4F0"
        fill_dark = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
        fill_blue = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        fill_light = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")
        right_al = Alignment(horizontal="right", vertical="center")
        left_al = Alignment(horizontal="left", vertical="center")
        thin = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )

        # Fila 1-2: título
        for rw in (1, 2):
            ws.merge_cells(start_row=rw, start_column=1, end_row=rw, end_column=10)
            for c in range(1, 11):
                ws.cell(row=rw, column=c).fill = fill_dark
        ws["A1"].value = "Resumen Vendedor por Proveedor"
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        ws["A1"].alignment = center
        ws["A2"].value = (
            f"Periodo: {fecha_ini} a {fecha_fin}  ·  Cifras en miles de pesos ($K)"
        )
        ws["A2"].font = Font(name="Calibri", size=10, color="FFFFFF")
        ws["A2"].alignment = center

        # Fila 4: encabezados
        headers = [
            "Zona", "Vendedor", "Proveedor", "Vta Neta ($K)",
            "Proyección ($K)", "% Cumpl.", "Cl. Impactados",
            "% Cl. Imp.", "Devolución ($K)", "% Dev.",
        ]
        hdr_font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = fill_blue
            cell.alignment = center
            cell.border = thin

        # Filas 5+: datos
        dat_font = Font(name="Calibri", size=9)
        prev_zona = None
        for ri, d in enumerate(data, 5):
            vals = [
                d["zona"], d["vendedor"], d["proveedor"],
                round(d["venta_neta"] / 1000), round(d["proyeccion"] / 1000),
                d["pct_cumpl"], d["cl_imp"], d["pct_cl"],
                round(d["devolucion"] / 1000), d["pct_dev"],
            ]
            is_new_zone = d["zona"] != prev_zona
            prev_zona = d["zona"]

            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = dat_font
                cell.border = thin
                if ci in (6, 8, 10):  # porcentajes
                    cell.number_format = "0.0%"
                elif ci in (4, 5, 7, 9):  # números
                    cell.number_format = "#,##0"
                cell.alignment = left_al if ci <= 3 else right_al
                if ri % 2 == 0:
                    cell.fill = fill_light

            # Separador visual entre zonas (borde superior grueso)
            if is_new_zone and ri > 5:
                for ci in range(1, 11):
                    c = ws.cell(row=ri, column=ci)
                    c.border = Border(
                        top=Side(style="medium", color=DARK),
                        left=thin.left, right=thin.right, bottom=thin.bottom,
                    )

        # Anchos
        widths = [8, 20, 25, 14, 16, 12, 14, 12, 16, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        _safe_save_workbook(wb, file_path)
        logger.info("Hoja Vendedor×Proveedor agregada: %d filas", len(data))

    except Exception as exc:
        logger.error("Error agregando hoja Vendedor×Proveedor: %s", exc, exc_info=True)


def _log_envio(engine, tipo, dest_id, dest_nombre, correos, fecha_ini, fecha_fin,
               archivo, estado, error_detalle=None):
    """Registra un envio en log_envio_reportes de la base BI."""
    from sqlalchemy import text as sa_text
    try:
        with engine.connect() as conn:
            conn.execute(
                sa_text(
                    "INSERT INTO log_envio_reportes "
                    "(tipo, destinatario_id, destinatario_nombre, correos, "
                    "fecha_ini, fecha_fin, archivo, estado, error_detalle) "
                    "VALUES (:tipo, :did, :dnombre, :correos, :fini, :ffin, "
                    ":archivo, :estado, :error)"
                ),
                {
                    "tipo": tipo, "did": dest_id, "dnombre": dest_nombre,
                    "correos": correos, "fini": fecha_ini, "ffin": fecha_fin,
                    "archivo": archivo, "estado": estado, "error": error_detalle,
                },
            )
            conn.commit()
    except Exception as exc:
        logger.warning("Error registrando log_envio_reportes: %s", exc)


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def enviar_reportes_email_task(database_name):
    """
    Genera y envia reportes por correo para UNA empresa.
    - Por cada proveedor activo: CuboVentas filtrado por proveedor_ids + Inventario filtrado por idProveedor
    - Por cada supervisor activo: CuboVentas filtrado por macrozonas + Inventario de todos los proveedores filtrado por bodegas
    El Excel resultante tiene dos hojas: Ventas e Inventario.
    Rango de ventas: 1ro del mes actual hasta hoy.
    """
    from sqlalchemy import text as sa_text
    from scripts.conexion import Conexion as Cnx
    from scripts.config import ConfigBasic as CB

    rq_job = get_current_job()
    job_id = rq_job.id if rq_job else None

    logger.info("[enviar_reportes_email_task] INICIO para %s", database_name)

    config = CB(database_name)
    c = config.config
    db_bi = c.get("dbBi")
    if not db_bi:
        return {"success": False, "error_message": f"No se encontro dbBi para {database_name}"}

    engine = Cnx.ConexionMariadb3(
        str(c.get("nmUsrIn")), str(c.get("txPassIn")),
        str(c.get("hostServerIn")), int(c.get("portServerIn")), db_bi,
    )

    # Rango de fechas: 1ro del mes actual -> hoy
    from datetime import date
    hoy = date.today()
    fecha_ini = hoy.replace(day=1).strftime("%Y-%m-%d")
    fecha_fin = hoy.strftime("%Y-%m-%d")

    enviados = 0
    errores = 0

    update_job_progress(job_id, 5, meta={"stage": "Leyendo destinatarios"})

    with engine.connect() as conn:
        # --- Proveedores ---
        proveedores = conn.execute(
            sa_text(
                "SELECT p.id, p.nombre, p.proveedor_ids, "
                "GROUP_CONCAT(pc.correo SEPARATOR ',') AS correos "
                "FROM proveedores_bi p "
                "JOIN proveedores_correo pc ON pc.proveedor_id = p.id AND pc.activo = 1 "
                "WHERE p.activo = 1 GROUP BY p.id"
            )
        ).mappings().all()

        # --- Supervisores ---
        supervisores = conn.execute(
            sa_text(
                "SELECT s.id, s.nombre, "
                "GROUP_CONCAT(DISTINCT sc.correo SEPARATOR ',') AS correos, "
                "GROUP_CONCAT(DISTINCT sm.macrozona_id SEPARATOR ',') AS macrozonas "
                "FROM supervisores s "
                "JOIN supervisores_correo sc ON sc.supervisor_id = s.id AND sc.activo = 1 "
                "LEFT JOIN supervisores_macrozona sm ON sm.supervisor_id = s.id "
                "WHERE s.activo = 1 GROUP BY s.id"
            )
        ).mappings().all()

    total = len(proveedores) + len(supervisores)
    if total == 0:
        return {"success": True, "message": "No hay destinatarios activos."}

    idx = 0

    # Enviar reportes a proveedores
    for prov in proveedores:
        idx += 1
        pct = int(10 + (idx / total) * 80)
        update_job_progress(job_id, pct, meta={"stage": f"Proveedor: {prov['nombre']}"})

        correos_list = [e.strip() for e in (prov["correos"] or "").split(",") if e.strip()]
        if not correos_list:
            continue

        try:
            # Generar reporte filtrado por proveedor_ids
            cubo = CuboVentas(
                database_name, fecha_ini, fecha_fin,
                user_id=None, reporte_id=7,
            )
            # Inyectar filtro de proveedor
            if prov["proveedor_ids"]:
                cubo.proveedores = [int(x.strip()) for x in prov["proveedor_ids"].split(",") if x.strip()]
            result = cubo.run()

            if result.get("success") and result.get("file_path"):
                # Agregar hoja de inventario filtrada por proveedor
                if prov["proveedor_ids"]:
                    prov_id_list = [int(x.strip()) for x in prov["proveedor_ids"].split(",") if x.strip()]
                    _add_inventario_sheet(result["file_path"], engine, proveedor_ids=prov_id_list)

                subject = f"Reporte {database_name} - {prov['nombre']} ({fecha_ini} a {fecha_fin})"
                body = (
                    f"<h3>Reporte de Ventas e Inventario - {prov['nombre']}</h3>"
                    f"<p>Periodo: {fecha_ini} a {fecha_fin}</p>"
                    f"<p>Empresa: {database_name}</p>"
                    f"<p>Adjunto encontrara el archivo Excel con el detalle de ventas e inventario.</p>"
                    f"<hr><small>Generado automaticamente por DataZenith.</small>"
                )
                _send_report_email(subject, correos_list, result["file_path"], body)
                _log_envio(engine, "proveedor", prov["id"], prov["nombre"],
                           prov["correos"], fecha_ini, fecha_fin,
                           result.get("file_path"), "enviado")
                enviados += 1
            else:
                _log_envio(engine, "proveedor", prov["id"], prov["nombre"],
                           prov["correos"], fecha_ini, fecha_fin, None,
                           "error", result.get("error_message", "Sin datos"))
                errores += 1
        except Exception as exc:
            logger.error("Error enviando reporte proveedor %s: %s", prov["nombre"], exc)
            _log_envio(engine, "proveedor", prov["id"], prov["nombre"],
                       prov["correos"], fecha_ini, fecha_fin, None, "error", str(exc))
            errores += 1

    # Enviar reportes a supervisores
    for sup in supervisores:
        idx += 1
        pct = int(10 + (idx / total) * 80)
        update_job_progress(job_id, pct, meta={"stage": f"Supervisor: {sup['nombre']}"})

        correos_list = [e.strip() for e in (sup["correos"] or "").split(",") if e.strip()]
        if not correos_list:
            continue

        try:
            cubo = CuboVentas(
                database_name, fecha_ini, fecha_fin,
                user_id=None, reporte_id=7,
            )
            # Inyectar filtro de macrozonas
            if sup["macrozonas"]:
                cubo.macrozonas = [int(x.strip()) for x in sup["macrozonas"].split(",") if x.strip()]
            result = cubo.run()

            if result.get("success") and result.get("file_path"):
                fp = result["file_path"]
                if sup["macrozonas"]:
                    macro_list = [int(x.strip()) for x in sup["macrozonas"].split(",") if x.strip()]
                    # Dashboard KPIs (se inserta como primera hoja)
                    _add_dashboard_supervisor_sheet(
                        fp, engine, macro_list, fecha_ini, fecha_fin, database_name,
                    )
                    # Vendedor × Proveedor
                    _add_vendedor_proveedor_sheet(
                        fp, engine, macro_list, fecha_ini, fecha_fin,
                    )
                    # Devoluciones del día
                    _add_devoluciones_dia_sheet(fp, engine, macro_list)
                    # Inventario
                    _add_inventario_sheet(fp, engine, macrozonas=macro_list)

                subject = f"Reporte {database_name} - {sup['nombre']} ({fecha_ini} a {fecha_fin})"
                body = (
                    f"<h3>Reporte Fuerza de Ventas - {sup['nombre']}</h3>"
                    f"<p>Periodo: {fecha_ini} a {fecha_fin}</p>"
                    f"<p>Empresa: {database_name}</p>"
                    f"<p>El archivo adjunto contiene:</p>"
                    f"<ul>"
                    f"<li><b>Dashboard</b> — KPIs por zona (ventas, devoluciones, cambios, drop size)</li>"
                    f"<li><b>Vendedor x Proveedor</b> — Venta, proyeccion, clientes, devoluciones por vendedor y proveedor</li>"
                    f"<li><b>Devoluciones del Dia</b> — Detalle por cliente</li>"
                    f"<li><b>Ventas</b> — Sabana de datos del mes</li>"
                    f"<li><b>Inventario</b> — Stock disponible por producto</li>"
                    f"</ul>"
                    f"<hr><small>Generado automaticamente por DataZenith.</small>"
                )
                _send_report_email(subject, correos_list, result["file_path"], body)
                _log_envio(engine, "supervisor", sup["id"], sup["nombre"],
                           sup["correos"], fecha_ini, fecha_fin,
                           result.get("file_path"), "enviado")
                enviados += 1
            else:
                _log_envio(engine, "supervisor", sup["id"], sup["nombre"],
                           sup["correos"], fecha_ini, fecha_fin, None,
                           "error", result.get("error_message", "Sin datos"))
                errores += 1
        except Exception as exc:
            logger.error("Error enviando reporte supervisor %s: %s", sup["nombre"], exc)
            _log_envio(engine, "supervisor", sup["id"], sup["nombre"],
                       sup["correos"], fecha_ini, fecha_fin, None, "error", str(exc))
            errores += 1

    update_job_progress(job_id, 100, meta={"stage": "Completado"})
    return {
        "success": errores == 0,
        "message": f"Enviados: {enviados}, Errores: {errores}",
        "enviados": enviados,
        "errores": errores,
    }


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def enviar_reporte_email_proveedor_task(database_name, proveedor_id):
    """
    Genera y envia el reporte por correo para UN proveedor especifico.
    Reutiliza la misma logica de enviar_reportes_email_task pero solo para un proveedor.
    """
    from sqlalchemy import text as sa_text
    from scripts.conexion import Conexion as Cnx
    from scripts.config import ConfigBasic as CB
    from datetime import date

    rq_job = get_current_job()
    job_id = rq_job.id if rq_job else None

    logger.info("[enviar_reporte_proveedor] INICIO para %s, proveedor_id=%s", database_name, proveedor_id)

    config = CB(database_name)
    c = config.config
    db_bi = c.get("dbBi")
    if not db_bi:
        return {"success": False, "error_message": f"No se encontro dbBi para {database_name}"}

    engine = Cnx.ConexionMariadb3(
        str(c.get("nmUsrIn")), str(c.get("txPassIn")),
        str(c.get("hostServerIn")), int(c.get("portServerIn")), db_bi,
    )

    hoy = date.today()
    fecha_ini = hoy.replace(day=1).strftime("%Y-%m-%d")
    fecha_fin = hoy.strftime("%Y-%m-%d")

    update_job_progress(job_id, 10, meta={"stage": "Leyendo proveedor"})

    with engine.connect() as conn:
        prov = conn.execute(
            sa_text(
                "SELECT p.id, p.nombre, p.proveedor_ids, "
                "GROUP_CONCAT(pc.correo SEPARATOR ',') AS correos "
                "FROM proveedores_bi p "
                "JOIN proveedores_correo pc ON pc.proveedor_id = p.id AND pc.activo = 1 "
                "WHERE p.id = :pid GROUP BY p.id"
            ),
            {"pid": proveedor_id},
        ).mappings().first()

    if not prov:
        return {"success": False, "error_message": f"Proveedor {proveedor_id} no encontrado o sin correos activos."}

    correos_list = [e.strip() for e in (prov["correos"] or "").split(",") if e.strip()]
    if not correos_list:
        return {"success": False, "error_message": f"El proveedor '{prov['nombre']}' no tiene correos activos."}

    update_job_progress(job_id, 30, meta={"stage": f"Generando reporte: {prov['nombre']}"})

    try:
        cubo = CuboVentas(
            database_name, fecha_ini, fecha_fin,
            user_id=None, reporte_id=7,
        )
        if prov["proveedor_ids"]:
            cubo.proveedores = [int(x.strip()) for x in prov["proveedor_ids"].split(",") if x.strip()]
        result = cubo.run()

        if result.get("success") and result.get("file_path"):
            update_job_progress(job_id, 70, meta={"stage": "Agregando inventario"})
            if prov["proveedor_ids"]:
                prov_id_list = [int(x.strip()) for x in prov["proveedor_ids"].split(",") if x.strip()]
                _add_inventario_sheet(result["file_path"], engine, proveedor_ids=prov_id_list)

            update_job_progress(job_id, 85, meta={"stage": "Enviando correo"})
            subject = f"Reporte {database_name} - {prov['nombre']} ({fecha_ini} a {fecha_fin})"
            body = (
                f"<h3>Reporte de Ventas e Inventario - {prov['nombre']}</h3>"
                f"<p>Periodo: {fecha_ini} a {fecha_fin}</p>"
                f"<p>Empresa: {database_name}</p>"
                f"<p>Adjunto encontrara el archivo Excel con el detalle de ventas e inventario.</p>"
                f"<hr><small>Generado automaticamente por DataZenith.</small>"
            )
            _send_report_email(subject, correos_list, result["file_path"], body)
            _log_envio(engine, "proveedor", prov["id"], prov["nombre"],
                       prov["correos"], fecha_ini, fecha_fin,
                       result.get("file_path"), "enviado")

            return {
                "success": True,
                "message": f"Reporte enviado a {prov['nombre']} ({', '.join(correos_list)})",
            }
        else:
            error_msg = result.get("error_message", "Sin datos para generar reporte")
            _log_envio(engine, "proveedor", prov["id"], prov["nombre"],
                       prov["correos"], fecha_ini, fecha_fin, None, "error", error_msg)
            return {"success": False, "error_message": error_msg}

    except Exception as exc:
        logger.error("Error enviando reporte proveedor %s: %s", prov["nombre"], exc, exc_info=True)
        _log_envio(engine, "proveedor", prov["id"], prov["nombre"],
                   prov["correos"], fecha_ini, fecha_fin, None, "error", str(exc))
        return {"success": False, "error_message": str(exc)}


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def enviar_reporte_email_supervisor_task(database_name, supervisor_id):
    """
    Genera y envia el reporte por correo para UN supervisor especifico.
    Incluye: Dashboard, Vendedor x Proveedor, Devoluciones del Dia, Ventas, Inventario.
    """
    from sqlalchemy import text as sa_text
    from scripts.conexion import Conexion as Cnx
    from scripts.config import ConfigBasic as CB
    from datetime import date

    rq_job = get_current_job()
    job_id = rq_job.id if rq_job else None

    logger.info("[enviar_reporte_supervisor] INICIO para %s, supervisor_id=%s", database_name, supervisor_id)

    config = CB(database_name)
    c = config.config
    db_bi = c.get("dbBi")
    if not db_bi:
        return {"success": False, "error_message": f"No se encontro dbBi para {database_name}"}

    engine = Cnx.ConexionMariadb3(
        str(c.get("nmUsrIn")), str(c.get("txPassIn")),
        str(c.get("hostServerIn")), int(c.get("portServerIn")), db_bi,
    )

    hoy = date.today()
    fecha_ini = hoy.replace(day=1).strftime("%Y-%m-%d")
    fecha_fin = hoy.strftime("%Y-%m-%d")

    update_job_progress(job_id, 10, meta={"stage": "Leyendo supervisor"})

    with engine.connect() as conn:
        sup = conn.execute(
            sa_text(
                "SELECT s.id, s.nombre, "
                "GROUP_CONCAT(DISTINCT sc.correo SEPARATOR ',') AS correos, "
                "GROUP_CONCAT(DISTINCT sm.macrozona_id SEPARATOR ',') AS macrozonas "
                "FROM supervisores s "
                "JOIN supervisores_correo sc ON sc.supervisor_id = s.id AND sc.activo = 1 "
                "LEFT JOIN supervisores_macrozona sm ON sm.supervisor_id = s.id "
                "WHERE s.id = :sid GROUP BY s.id"
            ),
            {"sid": supervisor_id},
        ).mappings().first()

    if not sup:
        return {"success": False, "error_message": f"Supervisor {supervisor_id} no encontrado o sin correos activos."}

    correos_list = [e.strip() for e in (sup["correos"] or "").split(",") if e.strip()]
    if not correos_list:
        return {"success": False, "error_message": f"El supervisor '{sup['nombre']}' no tiene correos activos."}

    update_job_progress(job_id, 20, meta={"stage": f"Generando reporte: {sup['nombre']}"})

    try:
        cubo = CuboVentas(
            database_name, fecha_ini, fecha_fin,
            user_id=None, reporte_id=7,
        )
        # Inyectar filtro de macrozonas
        if sup["macrozonas"]:
            cubo.macrozonas = [int(x.strip()) for x in sup["macrozonas"].split(",") if x.strip()]
        result = cubo.run()

        if result.get("success") and result.get("file_path"):
            fp = result["file_path"]
            update_job_progress(job_id, 50, meta={"stage": "Agregando hojas adicionales"})

            if sup["macrozonas"]:
                macro_list = [int(x.strip()) for x in sup["macrozonas"].split(",") if x.strip()]
                _add_dashboard_supervisor_sheet(
                    fp, engine, macro_list, fecha_ini, fecha_fin, database_name,
                )
                _add_vendedor_proveedor_sheet(
                    fp, engine, macro_list, fecha_ini, fecha_fin,
                )
                _add_devoluciones_dia_sheet(fp, engine, macro_list)
                _add_inventario_sheet(fp, engine, macrozonas=macro_list)

            update_job_progress(job_id, 85, meta={"stage": "Enviando correo"})
            subject = f"Reporte {database_name} - {sup['nombre']} ({fecha_ini} a {fecha_fin})"
            body = (
                f"<h3>Reporte Fuerza de Ventas - {sup['nombre']}</h3>"
                f"<p>Periodo: {fecha_ini} a {fecha_fin}</p>"
                f"<p>Empresa: {database_name}</p>"
                f"<p>El archivo adjunto contiene:</p>"
                f"<ul>"
                f"<li><b>Dashboard</b> — KPIs por zona (ventas, devoluciones, cambios, drop size)</li>"
                f"<li><b>Vendedor x Proveedor</b> — Venta, proyeccion, clientes, devoluciones por vendedor y proveedor</li>"
                f"<li><b>Devoluciones del Dia</b> — Detalle por cliente</li>"
                f"<li><b>Ventas</b> — Sabana de datos del mes</li>"
                f"<li><b>Inventario</b> — Stock disponible por producto</li>"
                f"</ul>"
                f"<hr><small>Generado automaticamente por DataZenith.</small>"
            )
            _send_report_email(subject, correos_list, fp, body)
            _log_envio(engine, "supervisor", sup["id"], sup["nombre"],
                       sup["correos"], fecha_ini, fecha_fin,
                       result.get("file_path"), "enviado")

            return {
                "success": True,
                "message": f"Reporte enviado a {sup['nombre']} ({', '.join(correos_list)})",
            }
        else:
            error_msg = result.get("error_message", "Sin datos para generar reporte")
            _log_envio(engine, "supervisor", sup["id"], sup["nombre"],
                       sup["correos"], fecha_ini, fecha_fin, None, "error", error_msg)
            return {"success": False, "error_message": error_msg}

    except Exception as exc:
        logger.error("Error enviando reporte supervisor %s: %s", sup["nombre"], exc, exc_info=True)
        _log_envio(engine, "supervisor", sup["id"], sup["nombre"],
                   sup["correos"], fecha_ini, fecha_fin, None, "error", str(exc))
        return {"success": False, "error_message": str(exc)}


# ── Wrappers para scheduling per-empresa (calculan fechas y ejecutan) ──

@job("default", timeout=DEFAULT_TIMEOUT)
def cdt_empresa_scheduled(empresa_id):
    """Nocturno CDT: ayer completo, una empresa."""
    from datetime import date, timedelta
    hoy = date.today()
    fi = (hoy - timedelta(days=1)).isoformat()
    ff = hoy.isoformat()
    planos_cdt_task(empresa_id, fi, ff, enviar_sftp=True)


@job("default", timeout=DEFAULT_TIMEOUT)
def tsol_empresa_scheduled(empresa_id):
    """Nocturno TSOL: mes anterior completo, una empresa."""
    from datetime import date, timedelta
    hoy = date.today()
    pm = hoy.replace(day=1) - timedelta(days=1)
    fi = pm.replace(day=1).isoformat()
    ff = pm.isoformat()
    planos_tsol_task(empresa_id, fi, ff, enviar_ftp=True)


@job("default", timeout=DEFAULT_TIMEOUT)
def cosmos_empresa_scheduled(empresa_id):
    """Nocturno Cosmos: 45 dias atras hasta hoy, una empresa."""
    from datetime import date, timedelta
    hoy = date.today()
    fi = (hoy - timedelta(days=45)).isoformat()
    ff = hoy.isoformat()
    planos_cosmos_task(empresa_id, fi, ff, enviar_ftps=True)


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def enviar_reportes_email_todas_empresas_task():
    """
    Orquestador: encola enviar_reportes_email_task para cada empresa
    que tenga envio_email_activo=True.
    """
    from apps.permisos.models import ConfEmpresas

    rq_job = get_current_job()
    job_id = rq_job.id if rq_job else None

    empresas = ConfEmpresas.objects.filter(estado=1, envio_email_activo=True)
    total = empresas.count()
    logger.info("[enviar_reportes_todas] Empresas con email activo: %d", total)

    if total == 0:
        return {"success": True, "message": "No hay empresas con envio de email activo."}

    encoladas = 0
    for i, emp in enumerate(empresas):
        update_job_progress(
            job_id, int((i / total) * 100),
            meta={"stage": f"Encolando {emp.name}"},
        )
        try:
            enviar_reportes_email_task.delay(emp.name)
            encoladas += 1
        except Exception as exc:
            logger.error("Error encolando email task para %s: %s", emp.name, exc)

    return {
        "success": True,
        "message": f"{encoladas}/{total} empresas encoladas para envio de reportes.",
        "encoladas": encoladas,
        "total": total,
    }


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def trazabilidad_task(
    database_name,
    IdtReporteIni,
    IdtReporteFin,
    user_id,
    report_id,
    batch_size=DEFAULT_BATCH_SIZE,
):
    """
    Tarea RQ para generar el reporte de Trazabilidad Preventa vs Facturacion.
    Lee datos de la tabla trazabilidad_preventa (BD BI) y genera Excel.
    """
    try:
        connection.close()
    except Exception:
        pass

    job = get_current_job()
    job_id = job.id if job else None
    logger.info(
        f"Iniciando trazabilidad_task (RQ Job ID: {job_id}) para DB: {database_name}, "
        f"Periodo: {IdtReporteIni}-{IdtReporteFin}"
    )

    def rq_update_progress(stage, progress_percent, current_rec=None, total_rec=None):
        meta = {"stage": stage}
        if current_rec is not None:
            meta["records_processed"] = current_rec
        if total_rec is not None:
            meta["total_records_estimate"] = total_rec
        update_job_progress(job_id, int(progress_percent), status="processing", meta=meta)

    extractor = TrazabilidadExtractor(
        database_name,
        IdtReporteIni,
        IdtReporteFin,
        user_id,
        progress_callback=rq_update_progress,
        batch_size=batch_size,
    )

    result_data = extractor.run()

    # Asegurar progreso final
    job = get_current_job()
    job_id = job.id if job else None
    if result_data.get("success"):
        update_job_progress(
            job_id, 100, status="completed",
            meta={"stage": "Completado", "file_ready": True},
        )
    else:
        update_job_progress(job_id, 100, status="failed", meta={"stage": "Fallido"})

    # Preview para el frontend
    if result_data.get("success"):
        try:
            preview = TrazabilidadExtractor.get_data(
                database_name, IdtReporteIni, IdtReporteFin, user_id,
                agrupacion="detalle", start=0, length=100,
            )
            result_data["preview_headers"] = preview.get("headers", [])
            result_data["preview_sample"] = preview.get("rows", [])
        except Exception as e:
            logger.warning(f"No se pudo obtener previsualización de trazabilidad: {e}")
            result_data["preview_headers"] = []
            result_data["preview_sample"] = []

    try:
        connection.close()
    except Exception:
        pass

    return result_data


# ══════════════════════════════════════════════════════════════════
# Tareas CDT (Planos para proveedores: MasterFoods, etc.)
# ══════════════════════════════════════════════════════════════════


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def planos_cdt_task(
    empresa_id,
    fecha_ini,
    fecha_fin,
    user_id=None,
    enviar_sftp=True,
):
    """
    Tarea RQ: Genera y envia planos CDT para una empresa.
    Extrae datos desde BD BI, genera archivos pipe-delimited y envia por SFTP.
    """
    try:
        connection.close()
    except Exception:
        pass

    from apps.permisos.models import CdtEnvio, ConfEmpresas

    job_obj = get_current_job()
    job_id = job_obj.id if job_obj else None

    logger.info(
        f"Iniciando planos_cdt_task (Job ID: {job_id}) "
        f"empresa={empresa_id}, periodo={fecha_ini} a {fecha_fin}"
    )

    update_job_progress(
        job_id, 5, "processing", meta={"stage": "Inicializando CDT"}
    )

    # Crear registro CdtEnvio
    empresa = ConfEmpresas.objects.get(id=empresa_id)

    envio = CdtEnvio.objects.create(
        empresa=empresa,
        fecha_inicio=fecha_ini,
        fecha_fin=fecha_fin,
        estado=CdtEnvio.Estado.PROCESANDO,
        usuario_id=user_id,
    )

    try:
        from scripts.cdt.PlanosCDT import PlanosCDT

        update_job_progress(
            job_id, 10, "processing", meta={"stage": "Extrayendo datos"}
        )

        processor = PlanosCDT(
            empresa_id=empresa_id,
            fecha_ini=fecha_ini,
            fecha_fin=fecha_fin,
            user_id=user_id,
            enviar_sftp=enviar_sftp,
        )

        update_job_progress(
            job_id, 30, "processing", meta={"stage": "Procesando planos"}
        )

        resultado = processor.procesar()

        # Actualizar registro CdtEnvio
        envio.estado = (
            CdtEnvio.Estado.ENVIADO
            if resultado.get("enviado_sftp")
            else CdtEnvio.Estado.PENDIENTE
        )
        envio.total_ventas = resultado.get("total_ventas", 0)
        envio.total_clientes = resultado.get("total_clientes", 0)
        envio.total_inventario = resultado.get("total_inventario", 0)
        envio.archivos_generados = json.dumps(
            resultado.get("archivos", []), ensure_ascii=False
        )
        envio.archivo_descarga = resultado.get("zip_path", "")
        envio.enviado_sftp = resultado.get("enviado_sftp", False)
        envio.log_ejecucion = resultado.get("log", "")
        envio.save()

        update_job_progress(
            job_id,
            100,
            "completed",
            meta={
                "stage": "Completado",
                "file_ready": bool(resultado.get("zip_path")),
            },
        )

        return {
            "success": True,
            "message": (
                f"Planos CDT generados: {len(resultado.get('archivos', []))} archivos"
            ),
            "file_path": resultado.get("zip_path"),
            "metadata": {
                "envio_id": envio.id,
                "total_ventas": resultado.get("total_ventas", 0),
                "total_clientes": resultado.get("total_clientes", 0),
                "total_inventario": resultado.get("total_inventario", 0),
                "archivos": resultado.get("archivos", []),
                "enviado_sftp": resultado.get("enviado_sftp", False),
                "stage": "Completado",
            },
        }

    except Exception as e:
        envio.estado = CdtEnvio.Estado.ERROR
        envio.log_ejecucion = str(e)
        envio.save()
        raise

    finally:
        try:
            connection.close()
        except Exception:
            pass


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def planos_cdt_todas_empresas_task():
    """
    Tarea nocturna: Ejecuta planos CDT para cada empresa con envio_cdt_activo=True.
    Cada empresa se procesa de forma independiente con su propio registro CdtEnvio.
    Se programa via django-rq-scheduler a las 11:00 PM Bogota.
    """
    try:
        connection.close()
    except Exception:
        pass

    from apps.permisos.models import ConfEmpresas, CdtEnvio
    from datetime import date, timedelta

    logger.info("Iniciando envio nocturno de planos CDT")

    # Periodo: ayer (un dia completo de ventas)
    hoy = date.today()
    fecha_ini = (hoy - timedelta(days=1)).isoformat()
    fecha_fin = hoy.isoformat()

    # Buscar empresas con CDT activo y configuracion CDT
    empresas = ConfEmpresas.objects.filter(
        envio_cdt_activo=True,
        estado=True,
        cdt_codigo_proveedor__isnull=False,
    ).exclude(cdt_codigo_proveedor="")

    resultados = []

    for empresa in empresas:
        logger.info(f"Procesando CDT: {empresa.name}")

        try:
            from scripts.cdt.PlanosCDT import PlanosCDT

            processor = PlanosCDT(
                empresa_id=empresa.id,
                fecha_ini=fecha_ini,
                fecha_fin=fecha_fin,
                enviar_sftp=True,
            )
            resultado = processor.procesar()

            CdtEnvio.objects.create(
                empresa=empresa,
                fecha_inicio=fecha_ini,
                fecha_fin=fecha_fin,
                estado=(
                    CdtEnvio.Estado.ENVIADO
                    if resultado.get("enviado_sftp")
                    else CdtEnvio.Estado.ERROR
                ),
                total_ventas=resultado.get("total_ventas", 0),
                total_clientes=resultado.get("total_clientes", 0),
                total_inventario=resultado.get("total_inventario", 0),
                archivos_generados=json.dumps(
                    resultado.get("archivos", []), ensure_ascii=False
                ),
                archivo_descarga=resultado.get("zip_path", ""),
                enviado_sftp=resultado.get("enviado_sftp", False),
                log_ejecucion=resultado.get("log", ""),
            )

            resultados.append({
                "empresa": empresa.name,
                "status": "ok",
                "archivos": len(resultado.get("archivos", [])),
            })

        except Exception as e:
            logger.error(
                f"Error procesando CDT para {empresa.name}: {e}"
            )
            resultados.append({
                "empresa": empresa.name,
                "status": "error",
                "error": str(e),
            })

    try:
        connection.close()
    except Exception:
        pass

    return {
        "success": True,
        "message": f"CDT nocturno: {len(resultados)} empresas procesadas",
        "metadata": {"resultados": resultados},
    }


# ══════════════════════════════════════════════════════════════════
# Tareas TSOL (Planos TrackSales para distribuidores)
# ══════════════════════════════════════════════════════════════════


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def planos_tsol_task(
    empresa_id,
    fecha_ini,
    fecha_fin,
    user_id=None,
    enviar_ftp=False,
):
    """
    Tarea RQ: Genera y envía planos TSOL para una empresa.
    Extrae datos desde BD BI, genera 11 archivos '{'-delimited y envía por FTP.
    """
    try:
        connection.close()
    except Exception:
        pass

    from apps.permisos.models import ConfEmpresas, TsolEnvio

    job_obj = get_current_job()
    job_id = job_obj.id if job_obj else None

    logger.info(
        f"Iniciando planos_tsol_task (Job ID: {job_id}) "
        f"empresa={empresa_id}, periodo={fecha_ini} a {fecha_fin}"
    )

    update_job_progress(
        job_id, 5, "processing", meta={"stage": "Inicializando TSOL"}
    )

    # Crear registro TsolEnvio
    empresa = ConfEmpresas.objects.get(id=empresa_id)

    envio = TsolEnvio.objects.create(
        empresa=empresa,
        fecha_inicio=fecha_ini,
        fecha_fin=fecha_fin,
        estado=TsolEnvio.Estado.PROCESANDO,
        usuario_id=user_id,
    )

    try:
        from scripts.tsol.PlanosTSOL import PlanosTSOL

        update_job_progress(
            job_id, 10, "processing", meta={"stage": "Extrayendo datos"}
        )

        processor = PlanosTSOL(
            empresa_id=empresa_id,
            fecha_ini=fecha_ini,
            fecha_fin=fecha_fin,
            user_id=user_id,
            enviar_ftp=enviar_ftp,
        )

        update_job_progress(
            job_id, 30, "processing", meta={"stage": "Procesando planos"}
        )

        resultado = processor.procesar()

        # Actualizar registro TsolEnvio
        totales = resultado.get("totales", {})
        envio.estado = (
            TsolEnvio.Estado.ENVIADO
            if resultado.get("enviado_ftp")
            else TsolEnvio.Estado.PENDIENTE
        )
        envio.total_ventas = totales.get("ventas", 0)
        envio.total_clientes = totales.get("clientes", 0)
        envio.total_productos = totales.get("productos", 0)
        envio.total_vendedores = totales.get("vendedores", 0)
        envio.total_inventario = totales.get("inventario", 0)
        envio.archivos_generados = json.dumps(
            resultado.get("archivos", []), ensure_ascii=False
        )
        envio.archivo_descarga = resultado.get("zip_path", "")
        envio.enviado_ftp = resultado.get("enviado_ftp", False)
        envio.log_ejecucion = resultado.get("log", "")
        envio.save()

        update_job_progress(
            job_id,
            100,
            "completed",
            meta={
                "stage": "Completado",
                "file_ready": bool(resultado.get("zip_path")),
            },
        )

        return {
            "success": True,
            "message": (
                f"Planos TSOL generados: {len(resultado.get('archivos', []))} archivos"
            ),
            "file_path": resultado.get("zip_path"),
            "metadata": {
                "envio_id": envio.id,
                "total_ventas": totales.get("ventas", 0),
                "total_clientes": totales.get("clientes", 0),
                "total_productos": totales.get("productos", 0),
                "total_vendedores": totales.get("vendedores", 0),
                "total_inventario": totales.get("inventario", 0),
                "archivos": resultado.get("archivos", []),
                "enviado_ftp": resultado.get("enviado_ftp", False),
                "stage": "Completado",
            },
        }

    except Exception as e:
        envio.estado = TsolEnvio.Estado.ERROR
        envio.log_ejecucion = str(e)
        envio.save()
        raise

    finally:
        try:
            connection.close()
        except Exception:
            pass


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def planos_tsol_todas_empresas_task():
    """
    Tarea nocturna: Ejecuta planos TSOL para cada empresa con envio_tsol_activo=True.
    Se programa via django-rq-scheduler.
    """
    try:
        connection.close()
    except Exception:
        pass

    from apps.permisos.models import ConfEmpresas, TsolEnvio
    from datetime import date, timedelta

    logger.info("Iniciando envío nocturno de planos TSOL")

    # Periodo: mes anterior completo
    hoy = date.today()
    primer_dia_mes = hoy.replace(day=1)
    ultimo_dia_mes_ant = primer_dia_mes - timedelta(days=1)
    primer_dia_mes_ant = ultimo_dia_mes_ant.replace(day=1)
    fecha_ini = primer_dia_mes_ant.isoformat()
    fecha_fin = ultimo_dia_mes_ant.isoformat()

    # Buscar empresas con TSOL activo y codigo configurado
    empresas = ConfEmpresas.objects.filter(
        envio_tsol_activo=True,
        estado=1,
        tsol_codigo__isnull=False,
    ).exclude(tsol_codigo="")

    resultados = []

    for empresa in empresas:
        logger.info(f"Procesando TSOL: {empresa.name} -> {empresa.tsol_nombre}")

        try:
            from scripts.tsol.PlanosTSOL import PlanosTSOL

            processor = PlanosTSOL(
                empresa_id=empresa.id,
                fecha_ini=fecha_ini,
                fecha_fin=fecha_fin,
                enviar_ftp=True,
            )
            resultado = processor.procesar()

            totales = resultado.get("totales", {})
            TsolEnvio.objects.create(
                empresa=empresa,
                fecha_inicio=fecha_ini,
                fecha_fin=fecha_fin,
                estado=(
                    TsolEnvio.Estado.ENVIADO
                    if resultado.get("enviado_ftp")
                    else TsolEnvio.Estado.ERROR
                ),
                total_ventas=totales.get("ventas", 0),
                total_clientes=totales.get("clientes", 0),
                total_productos=totales.get("productos", 0),
                total_vendedores=totales.get("vendedores", 0),
                total_inventario=totales.get("inventario", 0),
                archivos_generados=json.dumps(
                    resultado.get("archivos", []), ensure_ascii=False
                ),
                archivo_descarga=resultado.get("zip_path", ""),
                enviado_ftp=resultado.get("enviado_ftp", False),
                log_ejecucion=resultado.get("log", ""),
            )

            resultados.append({
                "empresa": empresa.name,
                "status": "ok",
                "archivos": len(resultado.get("archivos", [])),
            })

        except Exception as e:
            logger.error(
                f"Error procesando TSOL para {empresa.name}: {e}"
            )
            resultados.append({
                "empresa": empresa.name,
                "status": "error",
                "error": str(e),
            })

    try:
        connection.close()
    except Exception:
        pass

    return {
        "success": True,
        "message": f"TSOL nocturno: {len(resultados)} empresas procesadas",
        "metadata": {"resultados": resultados},
    }


# ══════════════════════════════════════════════════════════════════
# Tareas Cosmos (Planos para envío FTPS)
# ══════════════════════════════════════════════════════════════════


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def planos_cosmos_task(
    empresa_id,
    fecha_ini,
    fecha_fin,
    user_id=None,
    enviar_ftps=True,
):
    """
    Tarea RQ: Genera y envía planos Cosmos para una empresa.
    Extrae datos desde BD, genera archivos CSV pipe-delimited,
    comprime en ZIP y envía por FTPS.
    """
    try:
        connection.close()
    except Exception:
        pass

    from apps.permisos.models import CosmosEnvio, ConfEmpresas

    job_obj = get_current_job()
    job_id = job_obj.id if job_obj else None

    logger.info(
        f"Iniciando planos_cosmos_task (Job ID: {job_id}) "
        f"empresa={empresa_id}, periodo={fecha_ini} a {fecha_fin}"
    )

    update_job_progress(
        job_id, 5, "processing", meta={"stage": "Inicializando Cosmos"}
    )

    # Obtener empresa
    empresa = ConfEmpresas.objects.get(id=empresa_id)

    # Crear registro CosmosEnvio
    envio = CosmosEnvio.objects.create(
        empresa=empresa,
        fecha_inicio=fecha_ini,
        fecha_fin=fecha_fin,
        estado=CosmosEnvio.Estado.PROCESANDO,
        usuario_id=user_id,
    )

    try:
        import ast as ast_module
        from scripts.cosmos.planoscosmos import PlanosCosmos

        update_job_progress(
            job_id, 10, "processing", meta={"stage": "Extrayendo datos"}
        )

        # Parsear lista de IDs SQL Cosmos
        tx_cosmos = empresa.planos_cosmos
        if isinstance(tx_cosmos, str):
            try:
                tx_cosmos = ast_module.literal_eval(tx_cosmos)
            except (ValueError, SyntaxError):
                tx_cosmos = []

        # Configurar FTPS desde JSON
        cosmos_conn = empresa.cosmos_conexion or {}
        ftps_config = None
        if cosmos_conn.get("host"):
            ftps_config = {
                "host": cosmos_conn["host"],
                "port": cosmos_conn.get("port", 990),
                "user": cosmos_conn.get("user", ""),
                "pass": cosmos_conn.get("pass", ""),
                "remote_dir": cosmos_conn.get("ruta_remota", "/"),
                "certificate": cosmos_conn.get("certificate", ""),
            }

        # Directorio de salida bajo media/
        base_output_dir = os.path.join(
            "media", "cosmos", empresa.name.replace(" ", "_"),
        )

        processor = PlanosCosmos(
            database_name=empresa.name,
            empresa_id_cosmos=empresa.cosmos_empresa_id,
            fecha_ini=fecha_ini,
            fecha_fin=fecha_fin,
            tx_cosmos=tx_cosmos,
            ftps_config=ftps_config,
            enviar_ftps=enviar_ftps,
            base_output_dir=base_output_dir,
        )

        update_job_progress(
            job_id, 30, "processing", meta={"stage": "Generando archivos"}
        )

        resultado = processor.procesar_datos()

        update_job_progress(
            job_id, 80, "processing", meta={"stage": "Finalizando"}
        )

        # Actualizar registro CosmosEnvio
        envio.estado = (
            CosmosEnvio.Estado.ENVIADO
            if resultado.get("enviado_ftps")
            else CosmosEnvio.Estado.PENDIENTE
        )
        if not resultado.get("success"):
            envio.estado = CosmosEnvio.Estado.ERROR
        envio.total_registros = resultado.get("total_registros", 0)
        envio.archivos_generados = json.dumps(
            resultado.get("archivos", []), ensure_ascii=False
        )
        envio.archivo_descarga = resultado.get("zip_path", "")
        envio.enviado_ftps = resultado.get("enviado_ftps", False)
        envio.log_ejecucion = resultado.get("log", "")
        envio.save()

        update_job_progress(
            job_id,
            100,
            "completed",
            meta={
                "stage": "Completado",
                "file_ready": bool(resultado.get("zip_path")),
            },
        )

        return {
            "success": True,
            "message": (
                f"Planos Cosmos generados: {len(resultado.get('archivos', []))} archivos"
            ),
            "file_path": resultado.get("zip_path"),
            "metadata": {
                "envio_id": envio.id,
                "total_registros": resultado.get("total_registros", 0),
                "archivos": resultado.get("archivos", []),
                "enviado_ftps": resultado.get("enviado_ftps", False),
                "stage": "Completado",
            },
        }

    except Exception as e:
        envio.estado = CosmosEnvio.Estado.ERROR
        envio.log_ejecucion = str(e)
        envio.save()
        raise

    finally:
        try:
            connection.close()
        except Exception:
            pass


@job("default", timeout=DEFAULT_TIMEOUT, result_ttl=3600)
@task_handler
def planos_cosmos_todas_empresas_task():
    """
    Tarea nocturna: Ejecuta planos Cosmos para cada empresa con envio_cosmos_activo=True.
    """
    try:
        connection.close()
    except Exception:
        pass

    from apps.permisos.models import ConfEmpresas, CosmosEnvio
    from datetime import date, timedelta
    import ast as ast_module

    logger.info("Iniciando envío nocturno de planos Cosmos")

    # Periodo: 45 días atrás hasta hoy
    hoy = date.today()
    fecha_ini = (hoy - timedelta(days=45)).isoformat()
    fecha_fin = hoy.isoformat()

    empresas = ConfEmpresas.objects.filter(
        envio_cosmos_activo=True,
        estado=True,
        cosmos_empresa_id__isnull=False,
    ).exclude(cosmos_empresa_id="")

    resultados = []

    for empresa in empresas:
        logger.info(f"Procesando Cosmos: {empresa.name}")

        try:
            from scripts.cosmos.planoscosmos import PlanosCosmos

            tx_cosmos = empresa.planos_cosmos
            if isinstance(tx_cosmos, str):
                try:
                    tx_cosmos = ast_module.literal_eval(tx_cosmos)
                except (ValueError, SyntaxError):
                    tx_cosmos = []

            cosmos_conn = empresa.cosmos_conexion or {}
            ftps_config = None
            if cosmos_conn.get("host"):
                ftps_config = {
                    "host": cosmos_conn["host"],
                    "port": cosmos_conn.get("port", 990),
                    "user": cosmos_conn.get("user", ""),
                    "pass": cosmos_conn.get("pass", ""),
                    "remote_dir": cosmos_conn.get("ruta_remota", "/"),
                    "certificate": cosmos_conn.get("certificate", ""),
                }

            base_output_dir = os.path.join(
                "media", "cosmos", empresa.name.replace(" ", "_"),
            )

            processor = PlanosCosmos(
                database_name=empresa.name,
                empresa_id_cosmos=empresa.cosmos_empresa_id,
                fecha_ini=fecha_ini,
                fecha_fin=fecha_fin,
                tx_cosmos=tx_cosmos,
                ftps_config=ftps_config,
                enviar_ftps=True,
                base_output_dir=base_output_dir,
            )

            resultado = processor.procesar_datos()

            CosmosEnvio.objects.create(
                empresa=empresa,
                fecha_inicio=fecha_ini,
                fecha_fin=fecha_fin,
                estado=(
                    CosmosEnvio.Estado.ENVIADO
                    if resultado.get("enviado_ftps")
                    else CosmosEnvio.Estado.ERROR
                ),
                total_registros=resultado.get("total_registros", 0),
                archivos_generados=json.dumps(
                    resultado.get("archivos", []), ensure_ascii=False
                ),
                archivo_descarga=resultado.get("zip_path", ""),
                enviado_ftps=resultado.get("enviado_ftps", False),
                log_ejecucion=resultado.get("log", ""),
            )

            resultados.append({
                "empresa": empresa.name,
                "status": "ok",
                "archivos": len(resultado.get("archivos", [])),
            })

        except Exception as e:
            logger.error(
                f"Error procesando Cosmos para {empresa.name}: {e}"
            )
            resultados.append({
                "empresa": empresa.name,
                "status": "error",
                "error": str(e),
            })

    try:
        connection.close()
    except Exception:
        pass

    return {
        "success": True,
        "message": f"Cosmos nocturno: {len(resultados)} empresas procesadas",
        "metadata": {"resultados": resultados},
    }
